#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Wallpaper Engine (431960) 重复视频检测
- 严格只读（恢复 atime/mtime）
- 两阶段逻辑保留，但通过管线并行把“时长分桶”和“签名”重叠执行
- 中段取样（避免片头片尾黑屏）
- 临时缓存“匹完就删”（TemporaryDirectory 作用域内清理）
- 导出 XLSX：每组一行，组内条目按该条目命中的最大文件大小降序。
  同时导出两份：创意工坊链接版（取消订阅要用）和所在文件夹路径版（方便人工在资源管理器打开）。
- 进度条：整体仍有 S1/S2 两个 tqdm

现在的“重复判定”逻辑（与原版不同）：
  1. 先按“时长分桶 +（可选）音频指纹”做粗分桶；
  2. 在同一个粗桶内，使用 phash_parts 做模糊匹配：
     - 逐帧汉明距离归一化到 64 位基准
     - 组合分数 = 截尾均值 / (1 + 标准差)
     - 同内容不同分辨率：高标准差→分数低→容易匹配
     - 不同内容同模板：低标准差→分数不被压低→不易误匹配
     - 若分数 <= phash_distance_threshold（默认 1.5），就 union 成同一组
"""

import argparse
import contextlib
import hashlib
import io
import logging
import math
import os
import re
import shlex
import sqlite3
import subprocess
import sys
import tempfile
import time
import uuid
import json
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from functools import lru_cache  # NEW: 缓存 hex->ImageHash

try:
    import tomllib  # py311+
except Exception:
    import tomli as tomllib  # py310

from PIL import Image
import numpy as np
import imagehash
from openpyxl import Workbook
from openpyxl.styles import Font
from tqdm import tqdm  # 进度条

VIDEO_EXTS = {".mp4", ".mkv", ".webm", ".avi", ".mov", ".m4v", ".mpg", ".mpeg"}

# ----------------------------- 配置与数据 -----------------------------

@dataclass
class Config:
    workshop_root: str
    output_dir: str = "output"
    # 统一模型缓存根目录（HF / transformers / modelscope / torch.hub），支持相对路径。
    # 默认落在程序目录，便于用户整文件夹清理。
    model_cache_dir: str = "models_cache"
    # 筛重同时扫描 WE 本地项目（projects/myprojects 下各子文件夹内的视频）
    we_install_dir: str = ""
    include_myprojects: bool = True
    ffmpeg_path: str = "ffmpeg"
    ffprobe_path: str = "ffprobe"
    fpcalc_path: str = "fpcalc"

    # —— 中段取样设置 ——
    sample_frames: int = 36             # pHash 抽帧数量（在“视频窗口”内）
    phash_size: int = 12                # pHash 尺寸
    audio_window_seconds: int = 60      # 音频指纹“中间窗口”长度（秒）
    video_window_seconds: int = 15      # 视觉签名“中间窗口”长度（秒）
    seek_ratio: float = 0.5             # 窗口中心比例：0=开头，0.5=中点，1=结尾

    duration_rounding: str = "nearest_0.5"   # "int" 或 "nearest_0.5"
    require_both_signatures: bool = True
    # 跨相邻时长桶比较容差（秒）：>0 时，任意两个时长差 <= 此值的桶会被合并成一个临时桶
    # 跑一次 cluster_bucket_by_phash，用于救回"同源视频重编码 / ffprobe 估计误差 / 剪切"
    # 导致时长跨整数秒边界的对。0 关闭，回到原来的严格同桶比较。
    duration_cross_bucket_tolerance: float = 0.6

    # —— pHash 模糊匹配参数 ——
    # 针对以下业务需求校准：
    #   ✓ 合并：同视频不同码率 / 不同分辨率 / 带不带字幕
    #   ✗ 不合并：同视频不同角色 / 服饰 / 差分
    # 三种"要合并"场景里，"不同码率"的 pHash 距离最高，所以阈值按它的边缘情况设定。
    # "同模板不同角色"的双峰分布由 phash_bimodal_gap_cap 单独挡下。
    phash_distance_threshold: float = 1.5   # 组合距离分 = 截尾均值/(1+标准差)，越小越严格
    phash_trimmed_mean_cap: float = 12.0    # 截尾均值上限（64 位基准），> 此值直接判为不同内容
    phash_trim_ratio: float = 0.2           # 丢弃最高距离帧的比例，抑制对齐漂移帧
    phash_bimodal_gap_cap: float = 40.0     # 双峰差上限，> 此值即判为"同模板不同角色/服饰"类；放宽让语义闸把关

    # —— 颜色直方图（二次闸，专门挡"同场景不同角色/服饰"的 pHash 误报）——
    # pHash 是 DCT 低频特征，对"同构图+同动作+不同角色"会判相似；颜色直方图能捕捉服饰/肤色差异。
    # 流程：pHash 通过组合分判定"可能重复"后，再用 HSV 颜色直方图做二次确认。
    color_hist_bins_h: int = 16             # HSV 色相 bin 数
    color_hist_bins_s: int = 4              # HSV 饱和度 bin 数
    color_distance_threshold: float = 0.15  # Bhattacharyya 颜色距离阈值，> 此值认为角色/场景颜色不同

    # 音频软闸（作为 color 闸 rescue）：color 判拆时，若音频指纹几乎一致→豁免 color，继续走语义闸
    # 典型用途：同一视频但加了色调/滤镜版本（color_dist 超阈值），音频相同 → 应当合并
    audio_merge_override_color: bool = False  # 开关；默认关闭以保持兼容
    audio_merge_threshold: float = 0.15       # chromaprint 归一化汉明距离上限，< 此值视为音频相同

    # —— 视觉语义特征（可选三道闸，专治"同源视频局部差分"类）——
    # 例如：表情差分、道具差分、场景元素变体——少数帧有局部形变，pHash + 颜色直方图都会误判为同一视频。
    # 采取三个关键设计：
    #   1) 全片均匀抽 semantic_sample_frames 帧（而不是中段 15s），确保差分帧不漏；
    #   2) 按 index 对齐后逐帧算 cosine 距离（不是对平均 embedding 算距离）；
    #      pooled 策略会把少数差分帧稀释到 ~0.0001，无法区分。
    #   3) 三个指标并联（任一超过都判差分）：
    #      - mean  上限：整体性差异（短视频大比例帧都不同，mean 普遍 > 0.017）
    #      - max   上限：存在极端偏离帧（绝对值兜底）
    #      - peak_ratio = max/mean 上限：区分"水印/重编码全局漂移"（平坦, ratio≈2~3.5）
    #                   vs "同源局部差分"（尖峰, ratio≈4~6+）；
    #                   需要 max > peak_min_max 作为前置，防止极低 mean 时比率虚高。
    # 默认关闭；启用需要 pip 安装 torch（GPU 优先，CPU 也能跑但更慢）。
    # 首次启用会从 github 下载模型权重（dinov2_s ~84MB / dinov2_b ~330MB / dinov2_l ~1.1GB）。
    semantic_feature_enabled: bool = False
    semantic_feature_model: str = "dinov2_s"         # dinov2_s / dinov2_b / dinov2_l
    semantic_feature_device: str = "auto"            # auto / cuda / cpu
    semantic_sample_frames: int = 60                 # 语义专用：全片均匀抽多少帧（多于 pHash 的 36）
    # 三元判据（任一命中即判差分）——
    # 实测边界：MERGE（水印 / 不同码率）最高 mean=0.011 max=0.038 ratio=3.65；
    #         SPLIT（同源局部差分）最低 mean=0.005 max=0.019 ratio=3.94。
    semantic_distance_threshold: float = 0.015       # mean 上限
    semantic_max_threshold: float = 0.040            # max 上限
    semantic_peak_ratio_threshold: float = 3.8       # max/mean 上限（尖峰闸）
    semantic_peak_min_max: float = 0.015             # 尖峰闸前置：max 必须超过这个绝对值才启用
    # 编码漂移例外（两种模式，任一命中即豁免 max/ratio 尖峰闸；mean 闸不受影响）：
    #  A. 平坦漂移：p90 <= semantic_drift_p90_exempt，绝大多数帧几乎完全一致
    #  B. 稀疏超级尖峰：max > semantic_max_threshold 且"中间带" (0.5×max闸, max闸] 的帧数
    #     <= semantic_drift_sparse_mid_count 且 mean 小于等于 mean 闸阈值——
    #     刻画"压倒性一致 + 极少数帧飙很高但中间没有过渡"的分布，典型为水印孤立帧/decoder
    #     关键帧处理不一致；真差分的差异连续，中间带帧数普遍 >= 3。
    semantic_drift_p90_exempt: float = 0.005
    semantic_drift_sparse_mid_count: int = 2

    # —— Patch-level 空间分布闸（第四道闸，默认启用）——
    # 原理：DINOv2 ViT 输出 16×16 个 patch token，保留空间位置信息。两段视频按帧按空间位置
    # 做 patch 级 cosine，再看"高距 patch"聚在画面哪里：
    #   - 聚在边角 → 水印 / 字幕框（MERGE 强证据）
    #   - 聚在中心 → 主体差分 / 器官微调（SPLIT 强证据）
    # 作用范围：只在 ratio 闸触发、mean/max 闸未触发 的 borderline 对上启用，
    # 用来**推翻 ratio 闸对水印/重编码对的误判**；绝不改变 mean/max 闸判决——
    # 真差分（mean/max 超限）依旧照判 SPLIT，不受 patch 闸影响。
    semantic_patch_enabled: bool = True
    semantic_patch_grid: int = 8              # 16×16 patch 经 avgpool 到 grid×grid（默认 8×8=64 patches/帧）
    semantic_patch_hot_threshold: float = 0.015
    semantic_patch_min_hot_patches: int = 12
    semantic_patch_center_margin: float = 0.4   # 归一化距 <= 此值视为"中心"
    semantic_patch_edge_margin: float = 0.6     # 归一化距 >= 此值视为"边角"
    semantic_patch_corner_merge_frac: float = 0.55  # 边角热点占比 >= 此值 → 至少 weak 水印候选
    semantic_patch_center_split_frac: float = 0.45  # 中心热点占比 >= 此值 → content_diff（仅 log）
    # watermark_strong 阈值：推翻 max 闸 + ratio 闸（最强判据，真水印 PERS>=2 PC=1.00 typical）
    semantic_patch_persistent_frame_frac: float = 0.5   # >= 此比例帧都是热点 → "持久热点"
    semantic_patch_persistent_min: int = 2              # 持久热点数下限（真水印 PERS 2~7）
    semantic_patch_persistent_max: int = 8              # 持久热点数上限（超过视为稳定内容差异，非水印）
    semantic_patch_persistent_corner_min: float = 0.8   # 持久热点位于角落占比 >= 此值
    # watermark_weak 阈值：仅推翻 ratio 闸（轻度边缘抖动 / 低 max 尖峰）
    semantic_patch_weak_center_max: float = 0.12        # 中心热点占比 < 此值
    semantic_patch_weak_hot_ratio_max: float = 0.10     # 热点占全部 patch 比例 < 此值
                                                        # U4/U7/U8 0.05~0.09 vs G342/G344 0.10~0.12 的分界
    # content_diff_heavy：大面积稳定内容差异，即便 max/ratio 未触发也应 SPLIT（S44 型）
    semantic_patch_heavy_persistent_min: int = 10       # 持久热点数 >= 此值（真水印 <= 7）
    semantic_patch_heavy_hot_ratio_min: float = 0.20    # 热点占比 >= 此值（全图铺满）
    semantic_patch_heavy_pers_corner_max: float = 0.85  # 持久热点不全在角落（排除边角水印）
    semantic_patch_heavy_min_ratio: float = 2.5         # 语义 max/mean 比 >= 此值（防止平坦漂移型 P20 被误拆）
    # content_diff_center_persistent：微弱但稳定的中心差异（S43 型，语义距离很小但中心有持久不同）
    semantic_patch_center_persistent_corner_max: float = 0.20          # 持久热点几乎不在角落
    semantic_patch_center_persistent_total_corner_max: float = 0.25    # 总热点也不在角落主导
    # 即使 drift_exempt 命中（p90 极低的"编码漂移"例外），如果 patch 仍判定为
    # center_persistent 就挡下（S43 型：mean/max/ratio 都没触发但中心有微弱稳定差异）
    semantic_patch_drift_exempt_center_persistent_blocks: bool = True
    # watermark_anim：动画型水印（U6 型，位置抖动致 pers=0 但空间分布仍像水印）
    semantic_patch_anim_hot_ratio_min: float = 0.15    # 热点密度下限（U6=0.193；G342/G344/G346<0.12）
    semantic_patch_anim_corner_min: float = 0.65       # 角落热点占比下限（U6=0.78）
    semantic_patch_anim_center_max: float = 0.10       # 中心热点占比上限（U6=0.051；G340/G346>0.14）
    # watermark_anim 可推翻 max 闸，但上限保护：只有当 sd_max <= threshold × factor 时允许
    # （避免真大差异被误救）。2.5 → max <= 0.10，覆盖到 U6/M-anim 实测边界 0.08。
    semantic_patch_anim_max_override_factor: float = 2.5
    # ratio 闸 rescue：patch_verdict 虽然是 uncertain（没被判成 watermark_* 是因为
    # hot_ratio 或 center 恰好卡阈值），但空间分布已经是角落主导+中心干净+整体 mean 低，
    # 本质与水印一致。此路径只救 ratio 闸（不救 max 闸）。
    semantic_patch_ratio_rescue_corner_min: float = 0.60
    semantic_patch_ratio_rescue_center_max: float = 0.15
    semantic_patch_ratio_rescue_mean_max: float = 0.010
    # max 闸 rescue：比 ratio rescue 更危险（max 超标代表单帧极端差异），所以要求
    # 极度角落主导、中心极干净；max 仍受 anim_max_override_factor 上限保护。
    # 典型场景：1080p vs 4K 同源重编码导致角落压缩/锐化噪声 max 略超 0.04。
    semantic_patch_max_rescue_corner_min: float = 0.75
    semantic_patch_max_rescue_center_max: float = 0.10
    semantic_patch_max_rescue_mean_max: float = 0.010
    # —— 中心 patch mask 复验（纯像素 rescue 确认，可替代 LLM 文本仲裁）——
    # 当 ratio/max 闸 rescue 候选落入 uncertain 且满足角落主导/中心干净的空间条件后，
    # 在保留"内圈 keep × keep"格上重算 DINO patch cosine 距离。如果中心区域的 mean 与
    # max 都足够小，说明主要内容区域在两侧近乎一致——可直接放行合并，不依赖 LLM。
    # keep=4（对 8×8 grid 屏蔽外围两层，留 50% 居中面积）对 Wallpaper Engine 的角落/底部
    # 水印/字幕分布覆盖率 ≈ 95%。
    # 注意：阈值需收紧到比主 semantic_*_threshold 更严的水平（不然会误救"四角都有
    # 不同内容"的 G342/G344 型——那类对中心语义也巧合相似）。
    # 另强制 hot_ratio 极低（水印型的 hot 占比通常 < 0.06），避免真内容差异的对
    # 因为某几帧中心凑巧相似被误救。
    semantic_patch_center_mask_enabled: bool = True
    semantic_patch_center_mask_inner: int = 4
    semantic_patch_center_mask_mean_max: float = 0.006
    semantic_patch_center_mask_max_max: float = 0.025
    semantic_patch_center_mask_hot_ratio_max: float = 0.06
    # 二级放行层：只给 max-corner-dominant 的 uncertain 候选用。
    # 目标是救回"中心主体干净、但单帧有孤立峰值"的 1080p/4K 重编码对；因此除了放宽
    # center_max 外，还要求 center_p90 很低、角落更主导、中心更干净，且 dom_q 不高
    # （避免热点过度集中到单象限时误吞真内容差异）。
    semantic_patch_center_mask_relaxed_enabled: bool = True
    semantic_patch_center_mask_relaxed_hot_ratio_max: float = 0.14
    semantic_patch_center_mask_relaxed_dom_q_max: float = 0.30
    semantic_patch_center_mask_relaxed_corner_min: float = 0.85
    semantic_patch_center_mask_relaxed_center_max: float = 0.05
    semantic_patch_center_mask_relaxed_mean_max: float = 0.0065
    semantic_patch_center_mask_relaxed_p90_max: float = 0.010
    semantic_patch_center_mask_relaxed_max_max: float = 0.060

    # —— LLM 文本仲裁（可选，灰区双向 override）——
    # 作用定位：在像素判决落在阈值边缘时，读 project.json 的 title/description 给 LLM 做仲裁；
    # 专治"像素层看不出的差分变体"与"像素层挡下的同源再包装"（汉化/重压刚刚越线）。
    # 不启用时 0 依赖 0 显存（llm_arbitrator 模块懒加载）。
    llm_enabled: bool = False
    llm_model_name: str = "Qwen/Qwen3-8B"
    llm_device: str = "auto"                      # auto / cuda / cpu
    llm_quantization: str = "4bit"                # none / 4bit / 8bit
    llm_max_new_tokens: int = 32
    llm_max_input_tokens: int = 1024
    llm_enable_thinking: bool = False             # Qwen3 的 thinking mode；判决任务关掉更快
    # 灰区定义（SPLIT→MERGE 救回）：各闸挡下时，距离超出阈值不超过以下 margin 才调 LLM
    llm_gray_color_max_excess: float = 0.06
    llm_gray_semantic_mean_max_excess: float = 0.008
    llm_gray_semantic_max_max_excess: float = 0.02
    llm_gray_semantic_peak_max_excess: float = 1.5
    # MERGE→SPLIT 拦下：准合并前 patch_verdict 为 weak/降级 heavy 等可疑状态时调 LLM
    llm_intercept_on_patch_uncertain: bool = True
    # pixel-clean 短路：像素信号非常干净（mean & max 都低于下面两个阈值）时，
    # 即使 patch kind 是 uncertain / watermark_weak / content_diff，也不再叫 LLM 拦截，
    # 直接信任像素判决 MERGE。用来避免 LLM 仅因标题写法差异而把真同源对误拦。
    # 两个阈值需同时低于才算"干净"；只要一边超标仍走原来的 LLM 拦截流程。
    llm_intercept_pixel_clean_mean: float = 0.010
    llm_intercept_pixel_clean_max: float = 0.030
    # SPLIT→MERGE 救回：像素闸已判拆分，LLM 看文本决定要不要翻回合并。
    # 默认 False（"LLM 单向化"）——实测 rescue 方向风险大于收益（G117 AD 就被 rescue 误救），
    # 想启用请显式置 true。
    llm_rescue_enabled: bool = False
    # 跳过规则：两侧都没任何元数据时直接跳过（没信息可给 LLM）
    llm_skip_if_no_meta: bool = True

    # —— 人工覆写 ——
    # 算法无法区分的极少数边缘对（MERGE 与 SPLIT 在所有可量化特征上完全重叠），
    # 人工检查后，把 "wid_A|wid_B" 写进下面列表即可强制覆盖算法判决。
    # 顺序无关；重复对会自动去重；cluster 形成后再做 override。
    # force_merge_pairs：即便算法判拆，也强制合进同一组（把两侧 cluster union）
    # force_split_pairs：即便算法判合，也强制拆开（清除两侧所有跨组关系）
    force_merge_pairs: List[str] = field(default_factory=list)
    force_split_pairs: List[str] = field(default_factory=list)

    # —— 并行与超时 ——
    max_workers_stage1: int = 8         # 阶段1（测时长）线程数
    max_workers_stage2: int = 6         # 阶段2（pHash/音频）线程数
    ffprobe_timeout: int = 60
    ffmpeg_timeout: int = 60
    fpcalc_timeout: int = 60

    # —— 日志 / UI ——
    log_file: Optional[str] = None
    verbose: bool = False
    trace: bool = False
    progress: bool = True               # 显示进度条

@dataclass
class DurationRec:
    item_id: str
    path: Path
    size: int
    duration: Optional[float]
    bucket: Optional[str]
    url: str

@dataclass
class FileSig:
    item_id: str
    path: Path
    size: int
    duration: Optional[float]
    duration_bucket: Optional[str]
    phash_digest: Optional[str]
    phash_parts: List[str] = field(default_factory=list)
    audio_fp_digest: Optional[str] = None
    # Chromaprint 原始子指纹序列（uint32 逗号分隔字符串，来自 `fpcalc -raw`）。
    # 与 audio_fp_digest 并存：digest 用作硬分桶 key（SHA1），raw 用作软闸相似度比较。
    audio_fp_raw: Optional[str] = None
    color_histogram: Optional[List[float]] = None  # HSV 2D 直方图，平均到整个窗口后归一化
    # DINOv2 per-frame embedding 矩阵（N × D，L2 归一化）。用 numpy 存储，避免 List[List[float]] 的开销。
    semantic_embeddings: Optional[np.ndarray] = None
    # DINOv2 patch-level embedding（N × P × D，P=grid²，L2 归一化）。用于"空间分布闸"。
    # 缓存 schema 独立于 mean embedding（见 _cfg_semantic_patch_hash），可单独 bump 不影响上面的 mean 缓存。
    semantic_patch_embeddings: Optional[np.ndarray] = None
    # 文本元数据（可选，由 project.json 或文件名提供）；供 LLM 仲裁使用，不进缓存（每次扫描重读）。
    title_text: Optional[str] = None
    description_text: Optional[str] = None
    url: str = ""
    errors: List[str] = field(default_factory=list)

LOGGER = logging.getLogger("we_dup")

# ----------------------------- 日志 -----------------------------

def _utf8_stdout():
    """返回以 UTF-8 输出的流，避免 Windows 下文件名含 emoji 等字符时 gbk 编码报错。"""
    if hasattr(sys.stdout, "buffer"):
        return io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True
        )
    return sys.stdout


def setup_logging(level=logging.INFO, log_file: Optional[str] = None, file_mode: str = "a"):
    # 重要：先 close 旧 handler，避免 Windows 上文件句柄一直占用
    for h in list(LOGGER.handlers):
        try:
            h.close()
        except Exception:
            pass
    LOGGER.handlers.clear()
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    h = logging.StreamHandler(_utf8_stdout())
    h.setFormatter(fmt)
    LOGGER.addHandler(h)
    if log_file:
        fh = logging.FileHandler(log_file, encoding="utf-8", mode=file_mode)
        fh.setFormatter(fmt)
        LOGGER.addHandler(fh)
    LOGGER.setLevel(level)

# ----------------------------- 断点续跑缓存（sqlite） -----------------------------

def _cfg_visual_hash(cfg: "Config") -> str:
    data = {
        "sample_frames": int(cfg.sample_frames),
        "phash_size": int(cfg.phash_size),
        "video_window_seconds": float(cfg.video_window_seconds),
        "seek_ratio": float(cfg.seek_ratio),
        # 颜色直方图和 pHash 共用同一份抽帧结果；一起纳入视觉缓存版本，
        # bin 数变化 → 整个视觉签名（pHash + 颜色直方图）一起重算。
        "color_hist_bins_h": int(cfg.color_hist_bins_h),
        "color_hist_bins_s": int(cfg.color_hist_bins_s),
        # v2：改用 RGB 抽帧并附带颜色直方图；该常量用于强制让老 gray-only 缓存失效。
        "visual_schema_version": 2,
    }
    return hashlib.sha1(json.dumps(data, sort_keys=True).encode("utf-8")).hexdigest()


def _cfg_semantic_hash(cfg: "Config") -> str:
    """语义 embedding 的缓存 key。与 visual_cfg_hash 解耦：开/关/切模型不影响 pHash+颜色缓存。"""
    data = {
        "enabled": bool(cfg.semantic_feature_enabled),
        "model": str(cfg.semantic_feature_model),
        "semantic_sample_frames": int(cfg.semantic_sample_frames),
        # 语义抽帧固定是"全片均匀抽"，不跟 video_window_seconds / seek_ratio 挂钩
        # v2：从 pooled mean embedding 改为 per-frame embeddings（N×D 矩阵），存储格式变
        "semantic_schema_version": 2,
    }
    return hashlib.sha1(json.dumps(data, sort_keys=True).encode("utf-8")).hexdigest()


def _cfg_semantic_patch_hash(cfg: "Config") -> str:
    """Patch-level embedding 缓存 key。独立于 mean embedding 的 _cfg_semantic_hash：
    开/关 patch 闸、调整 grid 不会让原有 mean 缓存失效，反之亦然。
    """
    data = {
        "enabled": bool(cfg.semantic_patch_enabled),
        "model": str(cfg.semantic_feature_model),
        "semantic_sample_frames": int(cfg.semantic_sample_frames),
        "grid": int(cfg.semantic_patch_grid),
        "semantic_patch_schema_version": 1,
    }
    return hashlib.sha1(json.dumps(data, sort_keys=True).encode("utf-8")).hexdigest()


def _cfg_audio_hash(cfg: "Config") -> str:
    data = {
        "audio_window_seconds": int(cfg.audio_window_seconds),
        "seek_ratio": float(cfg.seek_ratio),
        "fpcalc_path": str(cfg.fpcalc_path),
        "ffmpeg_path": str(cfg.ffmpeg_path),
        # v2：保留原始 chromaprint 序列（audio_fp_raw）供软闸比较。
        # 老缓存只有 SHA1 digest，raw 为空，bump 版本强制重跑一次 fpcalc。
        "audio_schema_version": 2,
    }
    return hashlib.sha1(json.dumps(data, sort_keys=True).encode("utf-8")).hexdigest()


class SigCache:
    """
    sqlite 缓存：
    - duration：按文件 path+size+mtime_ns 缓存（跨运行复用）
    - signatures：视觉签名按 visual_cfg_hash；音频签名按 audio_cfg_hash
    """

    def __init__(self, db_path: Path):
        self.db_path = db_path
        self.lock = Lock()
        self._write_count = 0
        self._checkpoint_interval = 50
        self.conn = sqlite3.connect(str(db_path), timeout=30, check_same_thread=False)
        self.conn.execute("PRAGMA journal_mode=WAL;")
        self.conn.execute("PRAGMA synchronous=FULL;")
        with self.conn:
            self.conn.execute(
                """
                CREATE TABLE IF NOT EXISTS file_cache (
                  path TEXT PRIMARY KEY,
                  size INTEGER NOT NULL,
                  mtime_ns INTEGER NOT NULL,
                  item_id TEXT,
                  url TEXT,
                  duration REAL,
                  phash_digest TEXT,
                  phash_parts_json TEXT,
                  visual_cfg_hash TEXT,
                  audio_fp_digest TEXT,
                  audio_cfg_hash TEXT,
                  color_histogram_json TEXT,
                  semantic_embeddings_blob BLOB,
                  semantic_cfg_hash TEXT,
                  semantic_patch_blob BLOB,
                  semantic_patch_cfg_hash TEXT,
                  updated_ts INTEGER
                );
                """
            )
            # 旧库平滑升级：缺列时补上
            for col in ("semantic_embeddings_blob BLOB", "semantic_cfg_hash TEXT",
                        "semantic_patch_blob BLOB", "semantic_patch_cfg_hash TEXT",
                        "audio_fp_raw TEXT"):
                try:
                    self.conn.execute(f"ALTER TABLE file_cache ADD COLUMN {col}")
                except sqlite3.OperationalError:
                    pass  # 已存在

            # LLM 文本仲裁判决缓存：key = (sorted wid 对 | model_hash | prompt_hash | text_hash)。
            # 其中 text_hash 保证标题/描述变动后旧判决自动失效。
            self.conn.execute(
                """
                CREATE TABLE IF NOT EXISTS llm_judgement (
                  pair_key TEXT PRIMARY KEY,
                  wid_a TEXT NOT NULL,
                  wid_b TEXT NOT NULL,
                  model_name TEXT NOT NULL,
                  prompt_hash TEXT NOT NULL,
                  text_hash TEXT NOT NULL,
                  verdict TEXT NOT NULL,
                  reason TEXT,
                  raw TEXT,
                  updated_ts INTEGER
                );
                """
            )

    def _maybe_checkpoint(self):
        self._write_count += 1
        if self._write_count % self._checkpoint_interval == 0:
            try:
                self.conn.execute("PRAGMA wal_checkpoint(PASSIVE);")
            except Exception:
                pass

    def close(self):
        try:
            self.conn.execute("PRAGMA wal_checkpoint(TRUNCATE);")
        except Exception:
            pass
        try:
            self.conn.close()
        except Exception:
            pass

    @staticmethod
    def _key(p: Path) -> str:
        """
        Windows 文件系统大小写不敏感，而 Path.resolve() 返回的 UNC/盘符大小写依赖访问路径
        （`\\\\CS382\\...` vs `\\\\cs382\\...`），会被 SQLite 当成两个主键，造成同一视频重复 row、
        命中时拉到残缺的半份数据。统一 lower() 消除此抖动。
        """
        try:
            s = str(p.resolve())
        except Exception:
            s = str(p)
        if os.name == "nt":
            return s.lower()
        return s

    def delete_path(self, p: Path) -> None:
        """删除已不存在文件在缓存中的记录，避免库持续膨胀。"""
        key = self._key(p)
        with self.lock, self.conn:
            self.conn.execute("DELETE FROM file_cache WHERE path=?", (key,))

    def prune_missing(self) -> int:
        """删除缓存中路径已不存在的记录，返回删除条数。

        ⚠️ 不在启动时自动调用（见主流程）——保留已删文件的特征缓存，以便
           "取消订阅 → 文件删除 → 再次订阅/下载" 时无需重新计算特征。
        仅当你确实想瘦身数据库时才手动调用本方法。
        """
        with self.lock:
            cur = self.conn.execute("SELECT path FROM file_cache")
            rows = cur.fetchall()
        deleted = 0
        for (key,) in rows:
            try:
                if not Path(key).exists():
                    with self.lock, self.conn:
                        self.conn.execute("DELETE FROM file_cache WHERE path=?", (key,))
                    deleted += 1
            except Exception:
                pass
        return deleted

    def get_row(self, p: Path, size: int, mtime_ns: int) -> Optional[Dict[str, object]]:
        """
        命中条件：同路径、同 size。
        mtime 不再参与匹配——取消订阅后重新下载，文件字节内容相同但 mtime 改变，
        按 size 命中就可以直接复用已算好的 duration / pHash / 语义 / patch / 音频特征。
        视频内容确实改变时 size 几乎必然变化（bit-identical 巧合视为零概率）。
        """
        key = self._key(p)
        with self.lock:
            cur = self.conn.execute(
                "SELECT size,mtime_ns,item_id,url,duration,phash_digest,phash_parts_json,visual_cfg_hash,audio_fp_digest,audio_cfg_hash,color_histogram_json,semantic_embeddings_blob,semantic_cfg_hash,semantic_patch_blob,semantic_patch_cfg_hash,audio_fp_raw "
                "FROM file_cache WHERE path=?",
                (key,),
            )
            row = cur.fetchone()
        if not row:
            return None
        if int(row[0]) != int(size):
            return None
        return {
            "item_id": row[2],
            "url": row[3],
            "duration": row[4],
            "phash_digest": row[5],
            "phash_parts_json": row[6],
            "visual_cfg_hash": row[7],
            "audio_fp_digest": row[8],
            "audio_cfg_hash": row[9],
            "color_histogram_json": row[10],
            "semantic_embeddings_blob": row[11],
            "semantic_cfg_hash": row[12],
            "semantic_patch_blob": row[13],
            "semantic_patch_cfg_hash": row[14],
            "audio_fp_raw": row[15],
        }

    def upsert_duration(self, p: Path, size: int, mtime_ns: int, item_id: str, url: str, duration: Optional[float]):
        key = self._key(p)
        with self.lock, self.conn:
            self.conn.execute(
                """
                INSERT INTO file_cache(path,size,mtime_ns,item_id,url,duration,updated_ts)
                VALUES(?,?,?,?,?,?,?)
                ON CONFLICT(path) DO UPDATE SET
                  size=excluded.size,
                  mtime_ns=excluded.mtime_ns,
                  item_id=excluded.item_id,
                  url=excluded.url,
                  duration=excluded.duration,
                  updated_ts=excluded.updated_ts
                """,
                (key, int(size), int(mtime_ns), item_id, url, duration, int(time.time())),
            )
        self._maybe_checkpoint()

    def upsert_visual(self, p: Path, size: int, mtime_ns: int, item_id: str, url: str,
                      phash_digest: Optional[str], phash_parts: List[str],
                      color_histogram: Optional[List[float]],
                      visual_cfg_hash: str):
        key = self._key(p)
        parts_json = json.dumps(phash_parts, ensure_ascii=False)
        color_json = (
            json.dumps(color_histogram, ensure_ascii=False)
            if color_histogram is not None else None
        )
        with self.lock, self.conn:
            self.conn.execute(
                """
                INSERT INTO file_cache(path,size,mtime_ns,item_id,url,phash_digest,phash_parts_json,color_histogram_json,visual_cfg_hash,updated_ts)
                VALUES(?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(path) DO UPDATE SET
                  size=excluded.size,
                  mtime_ns=excluded.mtime_ns,
                  item_id=excluded.item_id,
                  url=excluded.url,
                  phash_digest=excluded.phash_digest,
                  phash_parts_json=excluded.phash_parts_json,
                  color_histogram_json=excluded.color_histogram_json,
                  visual_cfg_hash=excluded.visual_cfg_hash,
                  updated_ts=excluded.updated_ts
                """,
                (key, int(size), int(mtime_ns), item_id, url,
                 phash_digest, parts_json, color_json, visual_cfg_hash, int(time.time())),
            )
        self._maybe_checkpoint()

    def upsert_audio(self, p: Path, size: int, mtime_ns: int, item_id: str, url: str,
                     audio_fp_digest: Optional[str], audio_fp_raw: Optional[str],
                     audio_cfg_hash: str):
        key = self._key(p)
        with self.lock, self.conn:
            self.conn.execute(
                """
                INSERT INTO file_cache(path,size,mtime_ns,item_id,url,audio_fp_digest,audio_fp_raw,audio_cfg_hash,updated_ts)
                VALUES(?,?,?,?,?,?,?,?,?)
                ON CONFLICT(path) DO UPDATE SET
                  size=excluded.size,
                  mtime_ns=excluded.mtime_ns,
                  item_id=excluded.item_id,
                  url=excluded.url,
                  audio_fp_digest=excluded.audio_fp_digest,
                  audio_fp_raw=excluded.audio_fp_raw,
                  audio_cfg_hash=excluded.audio_cfg_hash,
                  updated_ts=excluded.updated_ts
                """,
                (key, int(size), int(mtime_ns), item_id, url,
                 audio_fp_digest, audio_fp_raw, audio_cfg_hash, int(time.time())),
            )
        self._maybe_checkpoint()

    def upsert_semantic(self, p: Path, size: int, mtime_ns: int, item_id: str, url: str,
                        semantic_embeddings: Optional[np.ndarray], semantic_cfg_hash: str):
        """
        写入 per-frame 语义 embedding 矩阵。格式 BLOB：
            header: uint32 N  +  uint32 D   (8 bytes, little-endian)
            body  : N × D × float16  （= N*D*2 bytes）
        与 visual 解耦：更换/关闭语义模型不会让 pHash/颜色缓存失效。
        """
        key = self._key(p)
        blob = pack_semantic_blob(semantic_embeddings)
        with self.lock, self.conn:
            self.conn.execute(
                """
                INSERT INTO file_cache(path,size,mtime_ns,item_id,url,semantic_embeddings_blob,semantic_cfg_hash,updated_ts)
                VALUES(?,?,?,?,?,?,?,?)
                ON CONFLICT(path) DO UPDATE SET
                  size=excluded.size,
                  mtime_ns=excluded.mtime_ns,
                  item_id=excluded.item_id,
                  url=excluded.url,
                  semantic_embeddings_blob=excluded.semantic_embeddings_blob,
                  semantic_cfg_hash=excluded.semantic_cfg_hash,
                  updated_ts=excluded.updated_ts
                """,
                (key, int(size), int(mtime_ns), item_id, url, blob, semantic_cfg_hash, int(time.time())),
            )
        self._maybe_checkpoint()

    def upsert_semantic_patch(self, p: Path, size: int, mtime_ns: int, item_id: str, url: str,
                              semantic_patch: Optional[np.ndarray], semantic_patch_cfg_hash: str):
        """
        写入 patch-level embedding 张量（N × P × D，float16）。
        独立于 semantic_embeddings_blob（mean 闸）；两套缓存 key 独立，互不失效。
        """
        key = self._key(p)
        blob = pack_semantic_patch_blob(semantic_patch)
        with self.lock, self.conn:
            self.conn.execute(
                """
                INSERT INTO file_cache(path,size,mtime_ns,item_id,url,semantic_patch_blob,semantic_patch_cfg_hash,updated_ts)
                VALUES(?,?,?,?,?,?,?,?)
                ON CONFLICT(path) DO UPDATE SET
                  size=excluded.size,
                  mtime_ns=excluded.mtime_ns,
                  item_id=excluded.item_id,
                  url=excluded.url,
                  semantic_patch_blob=excluded.semantic_patch_blob,
                  semantic_patch_cfg_hash=excluded.semantic_patch_cfg_hash,
                  updated_ts=excluded.updated_ts
                """,
                (key, int(size), int(mtime_ns), item_id, url, blob, semantic_patch_cfg_hash, int(time.time())),
            )
        self._maybe_checkpoint()

    # —— LLM 文本仲裁判决缓存 ——

    @staticmethod
    def _llm_pair_key(wid_a: str, wid_b: str, model_name: str,
                      prompt_hash: str, text_hash: str) -> str:
        lo, hi = sorted([str(wid_a), str(wid_b)])
        import hashlib as _h
        mh = _h.sha1(model_name.encode("utf-8")).hexdigest()[:8]
        return f"{lo}|{hi}|{mh}|{prompt_hash}|{text_hash}"

    def llm_get(self, wid_a: str, wid_b: str, model_name: str,
                prompt_hash: str, text_hash: str) -> Optional[Dict[str, str]]:
        key = self._llm_pair_key(wid_a, wid_b, model_name, prompt_hash, text_hash)
        with self.lock:
            cur = self.conn.execute(
                "SELECT verdict, reason, raw FROM llm_judgement WHERE pair_key=?",
                (key,),
            )
            row = cur.fetchone()
        if not row:
            return None
        return {"verdict": row[0], "reason": row[1] or "", "raw": row[2] or ""}

    def llm_upsert(self, wid_a: str, wid_b: str, model_name: str,
                   prompt_hash: str, text_hash: str,
                   verdict: str, reason: str, raw: str) -> None:
        key = self._llm_pair_key(wid_a, wid_b, model_name, prompt_hash, text_hash)
        lo, hi = sorted([str(wid_a), str(wid_b)])
        with self.lock, self.conn:
            self.conn.execute(
                """
                INSERT INTO llm_judgement(pair_key, wid_a, wid_b, model_name, prompt_hash,
                                           text_hash, verdict, reason, raw, updated_ts)
                VALUES(?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(pair_key) DO UPDATE SET
                  verdict=excluded.verdict,
                  reason=excluded.reason,
                  raw=excluded.raw,
                  updated_ts=excluded.updated_ts
                """,
                (key, lo, hi, model_name, prompt_hash, text_hash,
                 verdict, reason, raw, int(time.time())),
            )
        self._maybe_checkpoint()


# —— 语义 embedding BLOB 序列化（float16，N×D 矩阵）——

_SEMANTIC_BLOB_HEADER = 8  # 2 个 uint32 LE


def pack_semantic_blob(embs: Optional[np.ndarray]) -> Optional[bytes]:
    if embs is None:
        return None
    a = np.asarray(embs)
    if a.ndim != 2 or a.size == 0:
        return None
    n, d = a.shape
    header = np.array([n, d], dtype=np.uint32).tobytes()
    body = a.astype(np.float16, copy=False).tobytes()
    return header + body


def unpack_semantic_blob(blob: Optional[bytes]) -> Optional[np.ndarray]:
    if not blob or len(blob) < _SEMANTIC_BLOB_HEADER:
        return None
    try:
        n, d = np.frombuffer(blob[:_SEMANTIC_BLOB_HEADER], dtype=np.uint32).tolist()
        need = _SEMANTIC_BLOB_HEADER + int(n) * int(d) * 2
        if len(blob) != need or n == 0 or d == 0:
            return None
        arr = np.frombuffer(blob[_SEMANTIC_BLOB_HEADER:need], dtype=np.float16).reshape(int(n), int(d))
        return arr.astype(np.float32, copy=False)
    except Exception:
        return None


# —— patch-level embedding BLOB 序列化（float16，N × P × D 张量）——
#
# 头部格式：3 个 uint32 LE = 12 bytes
#   [N, P, D]
# 主体：N*P*D 个 float16（2 bytes each）
_PATCH_BLOB_HEADER = 12  # 3 × uint32 LE


def pack_semantic_patch_blob(arr: Optional[np.ndarray]) -> Optional[bytes]:
    if arr is None:
        return None
    a = np.asarray(arr)
    if a.ndim != 3 or a.size == 0:
        return None
    n, p, d = a.shape
    header = np.array([n, p, d], dtype=np.uint32).tobytes()
    body = a.astype(np.float16, copy=False).tobytes()
    return header + body


def unpack_semantic_patch_blob(blob: Optional[bytes]) -> Optional[np.ndarray]:
    if not blob or len(blob) < _PATCH_BLOB_HEADER:
        return None
    try:
        n, p, d = np.frombuffer(blob[:_PATCH_BLOB_HEADER], dtype=np.uint32).tolist()
        need = _PATCH_BLOB_HEADER + int(n) * int(p) * int(d) * 2
        if len(blob) != need or n == 0 or p == 0 or d == 0:
            return None
        arr = np.frombuffer(blob[_PATCH_BLOB_HEADER:need], dtype=np.float16).reshape(int(n), int(p), int(d))
        return arr.astype(np.float32, copy=False)
    except Exception:
        return None


_CACHE: Optional[SigCache] = None

# 语义特征模型单例（只在启用时加载；线程安全锁保护 embed 推理）
_SEMANTIC_MODEL: Optional[object] = None
_SEMANTIC_LOCK = Lock()

# LLM 文本仲裁模型：进程级单例 + 失败旗标。
# 放模块级而非 cluster_bucket_by_phash 闭包内是为了：
#   1. 多桶顺序跑同一次进程时，加载失败不会每桶重试一次刷屏日志；
#   2. LLM 模型首次加载开销大（5GB+ VRAM），多桶应复用同一个 handle。
_LLM_HANDLE: Optional[object] = None
_LLM_FAILED: bool = False
_LLM_LOCK = Lock()


def _ensure_semantic_model(cfg: "Config") -> Optional[object]:
    """按需加载语义模型单例。失败时返回 None 并在日志里说明原因；绝不抛出以免影响主流程。"""
    global _SEMANTIC_MODEL
    if not cfg.semantic_feature_enabled:
        return None
    if _SEMANTIC_MODEL is not None:
        return _SEMANTIC_MODEL
    with _SEMANTIC_LOCK:
        if _SEMANTIC_MODEL is not None:
            return _SEMANTIC_MODEL
        try:
            from semantic_features import load_semantic_model
            _SEMANTIC_MODEL = load_semantic_model(
                cfg.semantic_feature_model, cfg.semantic_feature_device,
                cache_dir=cfg.model_cache_dir,
            )
            return _SEMANTIC_MODEL
        except Exception as e:
            LOGGER.warning(
                "[semantic] 加载失败，将回退为仅 pHash+颜色闸：%s", e
            )
            _SEMANTIC_MODEL = None
            return None


# ----------------------------- 通用工具 -----------------------------

@contextlib.contextmanager
def preserve_times(p: Path):
    """保存并在退出时恢复 atime/mtime，保证最终不变。"""
    try:
        st = p.stat()
        at_ns, mt_ns = st.st_atime_ns, st.st_mtime_ns
    except Exception as e:
        LOGGER.debug(f"[times] stat 失败 {p}: {e}")
        at_ns = mt_ns = None
    try:
        yield
    finally:
        if at_ns is not None and mt_ns is not None:
            try:
                os.utime(p, ns=(at_ns, mt_ns), follow_symlinks=True)
                LOGGER.debug(f"[times] 恢复 atime/mtime：{p}")
            except Exception as e:
                LOGGER.warning(f"[times] 恢复 atime/mtime 失败 {p}: {e}")

def run_cmd(cmd: List[str], timeout: int, trace=False) -> Tuple[int, bytes, bytes, float]:
    """运行外部命令，返回 (rc, stdout, stderr, 耗时秒)。"""
    t0 = time.time()
    if trace:
        LOGGER.info("[exec] %s (timeout=%ss)", " ".join(shlex.quote(c) for c in cmd), timeout)
    try:
        flags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
        res = subprocess.run(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            timeout=timeout, check=False, creationflags=flags
        )
        el = time.time() - t0
        if trace:
            first_err = res.stderr.decode("utf-8", "ignore").splitlines()[:1]
            LOGGER.info("[exec] rc=%s in %.2fs%s", res.returncode, el,
                        f" | stderr: {first_err[0]}" if first_err else "")
        return res.returncode, res.stdout, res.stderr, el
    except subprocess.TimeoutExpired as e:
        el = time.time() - t0
        if trace:
            LOGGER.error("[exec] TIMEOUT after %.2fs: %s", el, " ".join(shlex.quote(c) for c in cmd))
        return 124, b"", str(e).encode("utf-8", "ignore"), el
    except Exception as e:
        el = time.time() - t0
        if trace:
            LOGGER.error("[exec] FAIL after %.2fs: %s", el, e)
        return 1, b"", str(e).encode("utf-8", "ignore"), el

def nearest_bucket(d: Optional[float], mode: str) -> Optional[str]:
    if d is None:
        return None
    m = (mode or "").strip().lower()
    if m in {"nearest_0.5", "0.5", "half"}:
        b = round(d * 2) / 2.0
    else:
        # 兼容历史写法：
        # - "int"：整数
        # - "nearest_1.0"：最接近 1 秒（等价于整数）
        # - 其他/空：回退到整数
        b = int(round(d))
    return str(b)

def make_we_url(item_id: str) -> str:
    return f"https://steamcommunity.com/sharedfiles/filedetails/?id={item_id}"


def make_item_url(item_id: str, video_path: Path) -> str:
    """workshop 项用 Steam 链接；myprojects 项（item_id 以 mp: 开头）用本地 file URI。"""
    if item_id.startswith("mp:"):
        try:
            return video_path.resolve().as_uri()
        except Exception:
            return str(video_path.resolve())
    return make_we_url(item_id)

def middle_window_start(duration: Optional[float], window: float, ratio: float) -> float:
    """计算以 ratio 为中心的窗口起点（ratio∈[0,1]；0=开头，0.5=中点，1=结尾）。"""
    if duration is None or duration <= 0:
        return 0.0
    window = min(window, duration)
    center = max(0.0, min(duration, duration * ratio))
    start = center - window / 2.0
    start = max(0.0, min(start, max(0.0, duration - window)))
    return float(start)

# ----------------------------- ffprobe / ffmpeg / fpcalc -----------------------------

def ffprobe_duration(ffprobe_path: str, video: Path, timeout: int, trace: bool) -> Optional[float]:
    with preserve_times(video):
        rc, out, err, _ = run_cmd(
            [ffprobe_path, "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", str(video)],
            timeout=timeout, trace=trace
        )
    if rc == 0:
        try:
            val = float(out.decode("utf-8", "ignore").strip())
            if math.isfinite(val) and val > 0:
                return val
        except Exception:
            pass
    with preserve_times(video):
        rc2, out2, err2, _ = run_cmd(
            [ffprobe_path, "-v", "error", "-show_entries", "stream=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", str(video)],
            timeout=timeout, trace=trace
        )
    if rc2 == 0:
        vals = []
        for line in out2.decode("utf-8", "ignore").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                f = float(line)
                if math.isfinite(f) and f > 0:
                    vals.append(f)
            except Exception:
                pass
        if vals:
            return max(vals)
    return None

def ffmpeg_extract_small_rgb_frames_middle(ffmpeg_path: str, video: Path, frames: int,
                                           timeout: int, trace: bool,
                                           duration: Optional[float],
                                           window_seconds: float,
                                           seek_ratio: float,
                                           size: int = 64) -> List[np.ndarray]:
    """
    在“中间窗口”内提取最多 N 张 `size`×`size` RGB 帧。

    默认 size=64 的用途（保持与 pHash / 颜色直方图兼容）：
      1. pHash（转灰度后算 DCT 感知哈希）
      2. 颜色直方图（在 HSV 空间统计色相/饱和度分布）
    size>=224 时用于语义模型推理（DINOv2 等）。

    重要：为了跨编码/跨分辨率的稳定性，这里**优先使用基于时间轴的均匀采样（fps=...）**，
    而不是优先 I 帧。原因是 I 帧位置强依赖编码器/GOP 设置，不同转码版本会抽到不同时间点，
    对“运动多/细节多”的 3D 视频（如 Blender 输出）会导致 pHash 大幅漂移，即使内容完全一致。
    """
    size = max(8, int(size))
    scale_expr = f"scale={size}:{size}"

    def run_and_collect(vf: str, limit: int, start_s: float, win_s: float) -> List[np.ndarray]:
        cmd = [
            ffmpeg_path, "-hide_banner", "-v", "error", "-nostdin",
            "-hwaccel", "auto",  # GPU 可用时用硬解码，不可用时自动退回 CPU
            "-ss", f"{start_s:.3f}",
            "-t",  f"{win_s:.3f}",
            "-i",  str(video),
            "-vf", vf,
            "-vsync", "vfr",
            "-frames:v", str(limit),
            "-f", "rawvideo", "-pix_fmt", "rgb24", "-"
        ]
        with preserve_times(video):
            rc, out, err, _ = run_cmd(cmd, timeout=timeout, trace=trace)
        if rc != 0 or not out:
            return []
        frame_size = size * size * 3
        n = len(out) // frame_size
        n = min(n, limit)
        buf = out[: n * frame_size]
        arr = np.frombuffer(buf, dtype=np.uint8).reshape((n, size, size, 3))
        return [arr[i] for i in range(n)]

    win = window_seconds if duration is None else min(window_seconds, duration)
    start = middle_window_start(duration, win, seek_ratio)

    fps = max(1.0, min(15.0, frames / max(1.0, win)))
    frames1 = run_and_collect(f"fps={fps},{scale_expr}", frames, start, win)
    if frames1:
        return frames1

    # 兜底：fps 抽帧失败时再尝试 I 帧（某些坏文件 fps 可能拿不到任何帧）
    frames2 = run_and_collect(f"select='eq(pict_type\\,I)',{scale_expr}", frames, start, win)
    if frames2:
        return frames2

    # 放宽窗口（×2，上限 60s）
    if duration and win < min(duration, 60.0):
        win2 = min(duration, min(60.0, win * 2.0))
        start2 = middle_window_start(duration, win2, seek_ratio)
        fps2 = max(1.0, min(15.0, frames / max(1.0, win2)))
        frames3 = run_and_collect(f"fps={fps2},{scale_expr}", frames, start2, win2)
        if frames3:
            return frames3

        frames4 = run_and_collect(f"select='eq(pict_type\\,I)',{scale_expr}", frames, start2, win2)
        if frames4:
            return frames4

    return []


def rgb_frames_to_gray(rgb_frames: List[np.ndarray]) -> List[np.ndarray]:
    """ITU-R BT.601 luma 近似：Y = 0.299R + 0.587G + 0.114B。"""
    out: List[np.ndarray] = []
    for rgb in rgb_frames:
        r = rgb[..., 0].astype(np.float32)
        g = rgb[..., 1].astype(np.float32)
        b = rgb[..., 2].astype(np.float32)
        y = (0.299 * r + 0.587 * g + 0.114 * b).clip(0, 255).astype(np.uint8)
        out.append(y)
    return out


def compute_color_histogram(rgb_frames: List[np.ndarray], bins_h: int, bins_s: int) -> Optional[List[float]]:
    """
    对多帧 RGB 求平均 HSV 2D 直方图。仅统计 H 和 S 两个通道（忽略 V=亮度，避免被编码/亮度曲线影响）。
    返回长度 bins_h*bins_s 的归一化浮点数组（sum=1），float 列表以便 JSON 序列化。
    """
    if not rgb_frames:
        return None
    hists: List[np.ndarray] = []
    for rgb in rgb_frames:
        try:
            img = Image.fromarray(rgb, "RGB").convert("HSV")
            hsv = np.asarray(img)
            h = hsv[..., 0].ravel().astype(np.int32)
            s = hsv[..., 1].ravel().astype(np.int32)
            hist2d, _, _ = np.histogram2d(
                h, s,
                bins=[bins_h, bins_s],
                range=[[0, 256], [0, 256]],
            )
            total = hist2d.sum()
            if total > 0:
                hist2d = hist2d / total
            hists.append(hist2d.flatten())
        except Exception as e:
            LOGGER.debug("[color_hist] 单帧计算失败: %s", e)
    if not hists:
        return None
    avg = np.mean(np.stack(hists, axis=0), axis=0)
    total = float(avg.sum())
    if total > 0:
        avg = avg / total
    return [float(x) for x in avg]


def color_histogram_distance(h1: Optional[List[float]], h2: Optional[List[float]]) -> float:
    """
    Bhattacharyya 距离：sqrt(1 - sum(sqrt(h1_i * h2_i)))。归一化到 [0, 1]，0=相同。
    任一侧缺失时返回 NaN-ish 值（float('nan')），由调用方决定如何处理。
    """
    if not h1 or not h2 or len(h1) != len(h2):
        return float("nan")
    a = np.asarray(h1, dtype=np.float64)
    b = np.asarray(h2, dtype=np.float64)
    a = np.clip(a, 0.0, None)
    b = np.clip(b, 0.0, None)
    bc = float(np.sum(np.sqrt(a * b)))
    bc = max(0.0, min(1.0, bc))
    return math.sqrt(max(0.0, 1.0 - bc))


_CHROMAPRINT_MIN_FRAMES = 40  # 至少 ~5 秒音频（fpcalc 约 8Hz 输出子指纹）


def _chromaprint_parse_raw(raw: Optional[str]) -> Optional[np.ndarray]:
    """将 `fpcalc -raw` 的 uint32 逗号序列解析成 np.ndarray[uint32]。"""
    if not raw:
        return None
    try:
        parts = raw.split(",")
        if len(parts) < _CHROMAPRINT_MIN_FRAMES:
            return None
        arr = np.fromiter((int(s) for s in parts if s), dtype=np.int64)
        if arr.size < _CHROMAPRINT_MIN_FRAMES:
            return None
        return (arr & 0xFFFFFFFF).astype(np.uint32)
    except Exception:
        return None


def chromaprint_distance(raw_a: Optional[str], raw_b: Optional[str]) -> Optional[float]:
    """计算两段 chromaprint 原始指纹的归一化汉明距离（∈[0,1]）。
    同源（bit-perfect 或轻微重编码）通常 < 0.05；同视频+不同 BGM 通常 > 0.30；
    不同视频（无特殊关系）> 0.40。
    任一侧缺失 / 太短 → 返回 None（调用方应视为"信号不足，abstain"）。
    对齐策略：取较短序列长度，从起点对齐（假定 `seek_ratio` 一致，起点大致相同）。
    """
    a = _chromaprint_parse_raw(raw_a)
    b = _chromaprint_parse_raw(raw_b)
    if a is None or b is None:
        return None
    n = int(min(a.size, b.size))
    if n < _CHROMAPRINT_MIN_FRAMES:
        return None
    xor = np.bitwise_xor(a[:n], b[:n])
    # np.unpackbits 需要 uint8 视图；把 uint32 数组展平成 uint8 再 unpack
    mismatch = int(np.unpackbits(xor.view(np.uint8)).sum())
    total_bits = n * 32
    return mismatch / float(total_bits) if total_bits > 0 else None

def compute_phash_from_frames(frames: List[np.ndarray], hash_size: int, prefix: str) -> Tuple[Optional[str], List[str]]:
    parts: List[str] = []
    for i, frame in enumerate(frames):
        try:
            img = Image.fromarray(frame)  # 不传 mode，避免 Pillow 弃用警告
            h = imagehash.phash(img, hash_size=hash_size)
            parts.append(str(h))
        except Exception as e:
            LOGGER.debug("%s pHash 帧 %d 失败: %s", prefix, i, e)
    if not parts:
        return None, []
    digest = hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()
    return digest, parts

def parse_fpcalc_stdout(stdout: bytes) -> Tuple[Optional[str], Optional[str]]:
    """返回 (digest_sha1, raw_fp_string)。
    - digest：FINGERPRINT 字符串的 SHA1，用作硬分桶 key（兼容老逻辑）
    - raw：原始 FINGERPRINT 字符串。`-raw` 模式为逗号分隔 uint32 序列，用于 chromaprint 汉明距离
    """
    text = stdout.decode("utf-8", "ignore")
    for line in text.splitlines():
        if line.startswith("FINGERPRINT="):
            fp = line.split("=", 1)[1].strip()
            if fp:
                dig = hashlib.sha1(fp.encode("utf-8")).hexdigest()
                return dig, fp
    return None, None

def fpcalc_on_wav(fpcalc_path: str, wav_path: Path, timeout: int, trace: bool
                  ) -> Tuple[Optional[str], Optional[str], str]:
    """对 WAV 用 `-raw` 模式（必选；软闸比较需要 uint32 序列，非 raw 模式输出是 base64 无法比较）。
    返回 (digest, raw_fp, reason)。
    """
    rc, out, err, _ = run_cmd([fpcalc_path, "-raw", str(wav_path)], timeout=timeout, trace=trace)
    if rc == 0:
        dig, raw = parse_fpcalc_stdout(out)
        if dig:
            return dig, raw, "ok"
    return None, None, f"rc={rc} {err.decode('utf-8','ignore').splitlines()[:1]}"

def fpcalc_fingerprint_middle(fpcalc_path: str, ffmpeg_path: str, video: Path,
                              duration: Optional[float],
                              audio_window_seconds: int,
                              seek_ratio: float,
                              timeouts: Tuple[int, int],
                              trace: bool) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    从视频“中间窗口”解 WAV（≤audio_window_seconds），再跑 fpcalc。
    使用 TemporaryDirectory + 随机文件名，并加 -y，确保用完即删且无“已存在”冲突。
    """
    win = audio_window_seconds if duration is None else min(audio_window_seconds, int(duration))
    start = middle_window_start(duration, win, seek_ratio)

    with tempfile.TemporaryDirectory(prefix="we_mid_") as tdir:
        wav_path = Path(tdir) / (uuid.uuid4().hex + ".wav")  # 文件尚不存在
        cmd_ff = [
            ffmpeg_path, "-hide_banner", "-v", "error", "-nostdin",
            "-hwaccel", "auto",  # 同样尝试硬解码
            "-ss", f"{start:.3f}",
            "-t",  f"{win:.3f}",
            "-i",  str(video),
            "-vn", "-ac", "1", "-ar", "11025",
            "-f",  "wav",
            "-y",  str(wav_path)  # 强制覆盖（以防极端残留）
        ]
        with preserve_times(video):
            rc1, out1, err1, _ = run_cmd(cmd_ff, timeout=timeouts[1], trace=trace)

        if rc1 != 0 or not wav_path.exists() or wav_path.stat().st_size == 0:
            return None, None, f"ffmpeg->wav rc={rc1} {err1.decode('utf-8','ignore').splitlines()[:1]}"

        dig, raw, reason = fpcalc_on_wav(fpcalc_path, wav_path, timeout=timeouts[0], trace=trace)
        if dig:
            return dig, raw, None
        return None, None, f"fpcalc on wav failed: {reason}"



# ----------------------------- 扫描 -----------------------------

def find_items(workshop_root: Path, myprojects_root: Optional[Path] = None) -> Dict[str, List[Path]]:
    """
    扫描创意工坊目录（仅数字 ID 子文件夹）与可选的 myprojects 目录。
    myprojects 下每个直接子文件夹为一项，item_id 统一为 mp:<文件夹名>，避免与 workshop 纯数字 ID 冲突。
    """
    items: Dict[str, List[Path]] = defaultdict(list)
    if not workshop_root.is_dir():
        raise FileNotFoundError(f"Workshop 路径不存在：{workshop_root}")
    for child in sorted(workshop_root.iterdir()):
        if not child.is_dir():
            continue
        item_id = child.name
        if not re.fullmatch(r"\d+", item_id):
            continue
        for root, _, files in os.walk(child):
            for fn in files:
                if Path(fn).suffix.lower() in VIDEO_EXTS:
                    items[item_id].append(Path(root) / fn)

    if myprojects_root is not None and myprojects_root.is_dir():
        for child in sorted(myprojects_root.iterdir()):
            if not child.is_dir():
                continue
            key = f"mp:{child.name}"
            for root, _, files in os.walk(child):
                for fn in files:
                    if Path(fn).suffix.lower() in VIDEO_EXTS:
                        items[key].append(Path(root) / fn)
    return items

# ----------------------------- 阶段1：测时长（保留旧实现，当前主流程不直接使用） -----------------------------

def measure_duration_one(item_id: str, vp: Path, cfg: Config) -> DurationRec:
    url = make_item_url(item_id, vp)
    if not vp.exists():
        # 文件缺失时仅跳过本轮筛重，保留缓存（便于取消订阅后再下载能直接命中）
        LOGGER.debug("[skip] 文件已删除，不参与筛重：%s", vp)
        return DurationRec(item_id=item_id, path=vp, size=0, duration=None, bucket=None, url=url)
    try:
        st = vp.stat()
        size = st.st_size
        mtime_ns = getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9))
    except Exception:
        size = 0
        mtime_ns = 0

    # 断点续跑：优先用缓存的 duration（仅当文件仍存在且未变），bucket 每次按当前配置重算
    if _CACHE and size and mtime_ns:
        row = _CACHE.get_row(vp, size, mtime_ns)
        if row and row.get("duration") is not None:
            if vp.exists():
                try:
                    st2 = vp.stat()
                    if st2.st_size == size and getattr(st2, "st_mtime_ns", int(st2.st_mtime * 1e9)) == mtime_ns:
                        dur = float(row["duration"]) if row["duration"] is not None else None
                        bucket = nearest_bucket(dur, cfg.duration_rounding)
                        return DurationRec(item_id=item_id, path=vp, size=size, duration=dur, bucket=bucket, url=url)
                except Exception:
                    pass

    dur = ffprobe_duration(cfg.ffprobe_path, vp, cfg.ffprobe_timeout, cfg.trace)
    bucket = nearest_bucket(dur, cfg.duration_rounding)
    if dur is None:
        LOGGER.warning("[dur] %s (%s) 时长获取失败", item_id, vp.name)

    if _CACHE and size and mtime_ns:
        try:
            _CACHE.upsert_duration(vp, size, mtime_ns, item_id, url, dur)
        except Exception as e:
            LOGGER.debug("[cache] upsert_duration failed: %s", e)
    return DurationRec(item_id=item_id, path=vp, size=size, duration=dur, bucket=bucket, url=url)

def stage1_measure_and_bucket(items_map: Dict[str, List[Path]], cfg: Config) -> Dict[str, List[DurationRec]]:
    """旧版：先把所有时长算完再返回候选桶（主流程现改为管线式，仅保留以备需要）"""
    total_files = sum(len(v) for v in items_map.values())
    LOGGER.info("[S1] 开始：并行测时长（max_workers=%d，files=%d）", cfg.max_workers_stage1, total_files)
    bucket_map: Dict[str, List[DurationRec]] = defaultdict(list)
    futures = []
    with ThreadPoolExecutor(max_workers=cfg.max_workers_stage1) as ex:
        for item_id, paths in items_map.items():
            for vp in paths:
                futures.append(ex.submit(measure_duration_one, item_id, vp, cfg))

        pb = tqdm(
            total=len(futures), desc="[S1] durations", unit="file",
            dynamic_ncols=True, disable=not cfg.progress,
            ascii=(os.name == "nt"),
        )
        try:
            for fut in as_completed(futures):
                rec = None
                try:
                    rec = fut.result()
                except Exception as e:
                    LOGGER.exception("[S1] 任务异常：%s", e)
                if rec and rec.duration is not None and rec.bucket is not None:
                    bucket_map[rec.bucket].append(rec)
                pb.update(1)
        finally:
            pb.close()

    candidate_buckets: Dict[str, List[DurationRec]] = {}
    for b, lst in bucket_map.items():
        item_ids = {r.item_id for r in lst}
        if len(item_ids) >= 2 and len(lst) >= 2:
            candidate_buckets[b] = lst

    LOGGER.info("[S1] 完成：总桶=%d，候选桶=%d（进入阶段2）", len(bucket_map), len(candidate_buckets))
    return candidate_buckets

# ----------------------------- 阶段2：签名（函数保留，主流程会在管线模式下调用） -----------------------------

def _read_meta_text(rec_path: Path, cfg: Config) -> Tuple[Optional[str], Optional[str]]:
    """只在 llm_enabled 时才读 project.json；缺失时降级用父目录+stem 作为 title。
    title/desc 都不进缓存（每次扫描重读，避免 project.json 变动后判决不更新）。
    """
    if not getattr(cfg, "llm_enabled", False):
        return None, None
    try:
        from llm_arbitrator import read_title_and_description, fallback_title_from_path
    except Exception:
        return None, None
    try:
        title, desc = read_title_and_description(rec_path)
        if not title and not desc:
            title = fallback_title_from_path(rec_path)
        return title, desc
    except Exception:
        return None, None


def sign_one(rec: DurationRec, cfg: Config) -> FileSig:
    """对单文件计算 pHash +（可选）音频指纹（均在中段窗口）。"""
    title_text, description_text = _read_meta_text(rec.path, cfg)
    if not rec.path.exists():
        # 文件缺失时仅跳过本轮筛重，保留缓存（便于取消订阅后再下载能直接命中）
        LOGGER.debug("[skip] 文件已删除，不参与筛重：%s", rec.path)
        return FileSig(
            item_id=rec.item_id, path=rec.path, size=rec.size,
            duration=rec.duration, duration_bucket=rec.bucket,
            phash_digest=None, phash_parts=[], audio_fp_digest=None, audio_fp_raw=None,
            title_text=title_text, description_text=description_text,
            url=rec.url,
            errors=["file_deleted"],
        )
    prefix = f"[{rec.item_id}]({rec.path.name})"
    LOGGER.info("%s 签名开始：%s (%.2f MiB)", prefix, str(rec.path), rec.size/1024/1024)

    # 尝试走缓存（仅当文件仍存在；视觉/音频按各自 cfg hash 命中）
    st = None
    try:
        st = rec.path.stat()
        mtime_ns = getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9))
    except Exception:
        mtime_ns = 0

    visual_hash = _cfg_visual_hash(cfg)
    audio_hash = _cfg_audio_hash(cfg)
    semantic_hash = _cfg_semantic_hash(cfg)
    semantic_patch_hash = _cfg_semantic_patch_hash(cfg)
    cached_phash_digest = None
    cached_phash_parts: List[str] = []
    cached_color_hist: Optional[List[float]] = None
    cached_audio = None
    cached_audio_raw: Optional[str] = None
    cached_semantic_embs: Optional[np.ndarray] = None
    cached_semantic_patch: Optional[np.ndarray] = None

    # 需要音频数据的两种情形：
    #   1) require_both_signatures：音频指纹参与硬分桶 key
    #   2) audio_merge_override_color：音频作为 color 闸的 rescue（软闸需要原始 fp）
    need_audio = bool(cfg.require_both_signatures) or bool(cfg.audio_merge_override_color)

    if _CACHE and rec.size and mtime_ns and rec.path.exists():
        row = _CACHE.get_row(rec.path, rec.size, mtime_ns)
        if row:
            if (row.get("visual_cfg_hash") == visual_hash
                    and row.get("phash_digest") and row.get("phash_parts_json")
                    and row.get("color_histogram_json")):
                cached_phash_digest = str(row.get("phash_digest") or "")
                try:
                    cached_phash_parts = json.loads(str(row.get("phash_parts_json") or "[]")) or []
                except Exception:
                    cached_phash_parts = []
                try:
                    ch_raw = json.loads(str(row.get("color_histogram_json") or "null"))
                    if isinstance(ch_raw, list) and ch_raw:
                        cached_color_hist = [float(x) for x in ch_raw]
                except Exception:
                    cached_color_hist = None
            if row.get("audio_cfg_hash") == audio_hash and row.get("audio_fp_digest"):
                cached_audio = str(row.get("audio_fp_digest") or "")
                raw_val = row.get("audio_fp_raw")
                if raw_val:
                    cached_audio_raw = str(raw_val)
            if (cfg.semantic_feature_enabled
                    and row.get("semantic_cfg_hash") == semantic_hash
                    and row.get("semantic_embeddings_blob")):
                cached_semantic_embs = unpack_semantic_blob(row.get("semantic_embeddings_blob"))
            if (cfg.semantic_patch_enabled
                    and row.get("semantic_patch_cfg_hash") == semantic_patch_hash
                    and row.get("semantic_patch_blob")):
                cached_semantic_patch = unpack_semantic_patch_blob(row.get("semantic_patch_blob"))

    visual_ok = bool(cached_phash_digest and cached_phash_parts and cached_color_hist)
    if not need_audio:
        audio_ok = True
    elif cfg.audio_merge_override_color:
        # 软闸比较需要 raw fp（老 v1 缓存没有）；没有 raw 就必须重跑
        audio_ok = bool(cached_audio) and bool(cached_audio_raw)
    else:
        audio_ok = bool(cached_audio)
    semantic_ok = (not cfg.semantic_feature_enabled) or cached_semantic_embs is not None
    semantic_patch_ok = (not cfg.semantic_patch_enabled) or cached_semantic_patch is not None

    if visual_ok and audio_ok and semantic_ok and semantic_patch_ok:
        if rec.path.exists():
            fs = FileSig(
                item_id=rec.item_id, path=rec.path, size=rec.size,
                duration=rec.duration, duration_bucket=rec.bucket,
                phash_digest=cached_phash_digest, phash_parts=cached_phash_parts,
                audio_fp_digest=cached_audio, audio_fp_raw=cached_audio_raw,
                color_histogram=cached_color_hist,
                semantic_embeddings=cached_semantic_embs,
                semantic_patch_embeddings=cached_semantic_patch,
                title_text=title_text, description_text=description_text,
                url=rec.url,
            )
            LOGGER.info("%s 命中缓存，跳过签名计算", prefix)
            LOGGER.info("%s 签名完成", prefix)
            return fs

    # pHash + 颜色直方图（中段窗口；共用一次抽帧结果）
    phash_digest = None
    phash_parts: List[str] = []
    color_hist: Optional[List[float]] = None
    rgb_frames = ffmpeg_extract_small_rgb_frames_middle(
        cfg.ffmpeg_path, rec.path, cfg.sample_frames, cfg.ffmpeg_timeout, cfg.trace,
        duration=rec.duration, window_seconds=float(cfg.video_window_seconds), seek_ratio=float(cfg.seek_ratio)
    )
    if not rgb_frames:
        LOGGER.warning("%s 帧提取失败（中间窗口）", prefix)
    else:
        gray_frames = rgb_frames_to_gray(rgb_frames)
        phash_digest, phash_parts = compute_phash_from_frames(gray_frames, cfg.phash_size, prefix)
        if not phash_digest:
            LOGGER.warning("%s pHash 计算失败", prefix)
        color_hist = compute_color_histogram(
            rgb_frames,
            bins_h=int(cfg.color_hist_bins_h),
            bins_s=int(cfg.color_hist_bins_s),
        )
        if color_hist is None:
            LOGGER.warning("%s 颜色直方图计算失败", prefix)

    if _CACHE and rec.size and mtime_ns and phash_parts:
        try:
            _CACHE.upsert_visual(rec.path, rec.size, mtime_ns, rec.item_id, rec.url,
                                 phash_digest, phash_parts, color_hist, visual_hash)
        except Exception as e:
            LOGGER.debug("[cache] upsert_visual failed: %s", e)

    # 语义 embedding（可选第三道闸；与 pHash/颜色独立缓存，可无代价开关）
    # + Patch-level embedding（第四道闸："空间分布"）。两者共用同一次抽帧，且共享模型。
    semantic_embs: Optional[np.ndarray] = cached_semantic_embs
    semantic_patch_embs: Optional[np.ndarray] = cached_semantic_patch
    need_mean = cfg.semantic_feature_enabled and semantic_embs is None
    need_patch = cfg.semantic_patch_enabled and semantic_patch_embs is None
    if need_mean or need_patch:
        sem_model = _ensure_semantic_model(cfg)
        if sem_model is not None:
            try:
                # 语义模型吃 224×224；独立抽一份，而且关键是 **全片均匀抽**（不是中段窗口）。
                # 原因：差分帧可能集中在视频某个段，中段窗口会漏掉。
                input_size = int(getattr(sem_model, "input_size", 224))
                full_window = float(rec.duration) if rec.duration and rec.duration > 0 else 9999.0
                sem_frames = ffmpeg_extract_small_rgb_frames_middle(
                    cfg.ffmpeg_path, rec.path, int(cfg.semantic_sample_frames),
                    cfg.ffmpeg_timeout, cfg.trace,
                    duration=rec.duration, window_seconds=full_window,
                    seek_ratio=0.0, size=input_size,
                )
                if sem_frames:
                    if need_mean:
                        with _SEMANTIC_LOCK:  # DINOv2 推理线程不安全
                            emb_mat = sem_model.embed_frames_per_frame(sem_frames)
                        if emb_mat is not None and emb_mat.size > 0:
                            semantic_embs = emb_mat.astype(np.float32, copy=False)
                            if _CACHE and rec.size and mtime_ns:
                                try:
                                    _CACHE.upsert_semantic(
                                        rec.path, rec.size, mtime_ns, rec.item_id, rec.url,
                                        semantic_embs, semantic_hash,
                                    )
                                except Exception as e:
                                    LOGGER.debug("[cache] upsert_semantic failed: %s", e)
                    if need_patch:
                        try:
                            with _SEMANTIC_LOCK:
                                patch_mat = sem_model.embed_frames_patch_grid(
                                    sem_frames, grid_side=int(cfg.semantic_patch_grid)
                                )
                        except Exception as e:
                            LOGGER.warning("%s patch-level embedding 失败：%s", prefix, e)
                            patch_mat = None
                        if patch_mat is not None and patch_mat.size > 0:
                            semantic_patch_embs = patch_mat.astype(np.float32, copy=False)
                            if _CACHE and rec.size and mtime_ns:
                                try:
                                    _CACHE.upsert_semantic_patch(
                                        rec.path, rec.size, mtime_ns, rec.item_id, rec.url,
                                        semantic_patch_embs, semantic_patch_hash,
                                    )
                                except Exception as e:
                                    LOGGER.debug("[cache] upsert_semantic_patch failed: %s", e)
                else:
                    LOGGER.warning("%s 语义抽帧失败（%dx%d）", prefix, input_size, input_size)
            except Exception as e:
                LOGGER.warning("%s 语义 embedding 计算失败：%s", prefix, e)

    # 音频指纹（中段窗口）——当 require_both_signatures 或 audio_merge_override_color 开启时才做
    audio_digest: Optional[str] = cached_audio
    audio_raw: Optional[str] = cached_audio_raw
    if need_audio and not audio_ok:
        ad, araw, rsn = fpcalc_fingerprint_middle(
            cfg.fpcalc_path, cfg.ffmpeg_path, rec.path, duration=rec.duration,
            audio_window_seconds=int(cfg.audio_window_seconds),
            seek_ratio=float(cfg.seek_ratio),
            timeouts=(cfg.fpcalc_timeout, cfg.ffmpeg_timeout),
            trace=cfg.trace
        )
        audio_digest = ad
        audio_raw = araw
        if not audio_digest:
            LOGGER.warning("%s 音频指纹获取失败（%s）", prefix, rsn)
        if _CACHE and rec.size and mtime_ns:
            try:
                _CACHE.upsert_audio(
                    rec.path, rec.size, mtime_ns, rec.item_id, rec.url,
                    audio_digest, audio_raw, audio_hash,
                )
            except Exception as e:
                LOGGER.debug("[cache] upsert_audio failed: %s", e)

    fs = FileSig(
        item_id=rec.item_id, path=rec.path, size=rec.size,
        duration=rec.duration, duration_bucket=rec.bucket,
        phash_digest=phash_digest, phash_parts=phash_parts,
        audio_fp_digest=audio_digest, audio_fp_raw=audio_raw,
        color_histogram=color_hist,
        semantic_embeddings=semantic_embs,
        semantic_patch_embeddings=semantic_patch_embs,
        title_text=title_text, description_text=description_text,
        url=rec.url,
    )
    if cfg.require_both_signatures:
        if not (rec.bucket and phash_digest and audio_digest):
            fs.errors.append("incomplete_signature")
    else:
        if not (rec.bucket and phash_digest):
            fs.errors.append("no_visual_signature")

    LOGGER.info("%s 签名完成", prefix)
    return fs

def stage2_sign_candidates(candidate_buckets: Dict[str, List[DurationRec]], cfg: Config) -> List[FileSig]:
    """旧版：拿到完整候选桶后再统一并行签名（主流程现改为管线式，仅保留以备需要）"""
    total_candidates = sum(len(v) for v in candidate_buckets.values())
    LOGGER.info("[S2] 开始：候选桶签名（max_workers=%d，files=%d）", cfg.max_workers_stage2, total_candidates)
    filesigs: List[FileSig] = []
    futures = []
    with ThreadPoolExecutor(max_workers=cfg.max_workers_stage2) as ex:
        for bucket, recs in candidate_buckets.items():
            for rec in recs:
                futures.append(ex.submit(sign_one, rec, cfg))

        pb = tqdm(
            total=len(futures), desc="[S2] signatures", unit="file",
            dynamic_ncols=True, disable=not cfg.progress,
            ascii=(os.name == "nt"),
        )
        try:
            for fut in as_completed(futures):
                try:
                    fs = fut.result()
                    filesigs.append(fs)
                except Exception as e:
                    LOGGER.exception("[S2] 签名任务异常：%s", e)
                pb.update(1)
        finally:
            pb.close()

    LOGGER.info("[S2] 完成：已计算签名文件数=%d", len(filesigs))
    return filesigs

# ----------------------------- 新：阶段1+阶段2 管线并行 -----------------------------

def stage1_and_stage2_pipelined(items_map: Dict[str, List[Path]], cfg: Config) -> List[FileSig]:
    """
    管线式执行：
      - 线程池 A：测时长 + 分桶
      - 线程池 B：一旦某个桶变成候选桶，就立刻对该桶内文件提交签名任务
      - 这样 ffprobe / ffmpeg / fpcalc 可以同时跑，多核 / GPU 利用率更高
    """
    total_files = sum(len(v) for v in items_map.values())
    LOGGER.info("[PIPE] 开始：阶段1+阶段2 管线并行（dur_workers=%d, sig_workers=%d, files=%d）",
                cfg.max_workers_stage1, cfg.max_workers_stage2, total_files)

    bucket_map: Dict[str, List[DurationRec]] = defaultdict(list)
    bucket_items: Dict[str, set] = defaultdict(set)
    candidate_buckets: set = set()
    sig_futures = []
    filesigs: List[FileSig] = []
    lock = Lock()

    with ThreadPoolExecutor(max_workers=cfg.max_workers_stage1) as ex_dur, \
         ThreadPoolExecutor(max_workers=cfg.max_workers_stage2) as ex_sig:

        # 提交所有“测时长”任务
        dur_futs = []
        for item_id, paths in items_map.items():
            for vp in paths:
                dur_futs.append(ex_dur.submit(measure_duration_one, item_id, vp, cfg))

        pb1 = tqdm(
            total=len(dur_futs), desc="[S1] durations", unit="file",
            dynamic_ncols=True, disable=not cfg.progress,
            ascii=(os.name == "nt"),
        )
        try:
            for fut in as_completed(dur_futs):
                rec = None
                try:
                    rec = fut.result()
                except Exception as e:
                    LOGGER.exception("[PIPE-S1] 任务异常：%s", e)

                if rec and rec.duration is not None and rec.bucket is not None:
                    to_sign: List[DurationRec] = []
                    # 更新桶信息 & 判断是否成为候选桶
                    with lock:
                        bucket_map[rec.bucket].append(rec)
                        items = bucket_items[rec.bucket]
                        if rec.item_id not in items:
                            items.add(rec.item_id)
                        if rec.bucket not in candidate_buckets:
                            # 第一次成为候选桶：这个桶里所有已有文件都要签名
                            if len(bucket_map[rec.bucket]) >= 2 and len(items) >= 2:
                                candidate_buckets.add(rec.bucket)
                                to_sign = list(bucket_map[rec.bucket])
                        else:
                            # 已经是候选桶了，新文件直接签名
                            to_sign = [rec]

                    # 在锁外提交签名任务，避免阻塞其他时长任务
                    for r2 in to_sign:
                        sig_futures.append(ex_sig.submit(sign_one, r2, cfg))

                pb1.update(1)
        finally:
            pb1.close()

        LOGGER.info("[PIPE] 阶段1结束，候选桶=%d，已触发签名任务=%d", len(candidate_buckets), len(sig_futures))

        # 等待所有签名任务完成（此时阶段2可能已经完成了一部分）
        pb2 = tqdm(
            total=len(sig_futures), desc="[S2] signatures", unit="file",
            dynamic_ncols=True, disable=not cfg.progress,
            ascii=(os.name == "nt"),
        )
        try:
            for sf in as_completed(sig_futures):
                try:
                    fs = sf.result()
                    filesigs.append(fs)
                except Exception as e:
                    LOGGER.exception("[PIPE-S2] 签名任务异常：%s", e)
                pb2.update(1)
        finally:
            pb2.close()

    LOGGER.info("[PIPE] 完成：签名文件数=%d", len(filesigs))
    return filesigs

# ----------------------------- pHash 模糊匹配工具 -----------------------------

@lru_cache(maxsize=8192)
def _hex_to_hash_cached(h: str) -> imagehash.ImageHash:
    """缓存 hex->ImageHash，避免重复解析带来的开销。"""
    return imagehash.hex_to_hash(h)

def phash_distance(fs1: FileSig, fs2: FileSig,
                   tm_cap: float = 4.0,
                   trim_ratio: float = 0.1,
                   bimodal_gap_cap: float = 15.0) -> float:
    """
    组合分数 = 截尾均值 / (1 + 标准差)，附带两道排除闸。

    逐帧距离归一化到 8x8（64 位）基准后排序，然后：

    1. 双峰差闸（核心判据）：
       "高半段均值 - 低半段均值" > `bimodal_gap_cap` 即判为不同内容。
       - 同视频不同编码：单峰分布，差通常 < 12
       - 同模板不同角色/服装：双峰（背景帧≈0 + 角色帧≈25），差常 ≥ 18
       这是区分这两种场景最鲁棒的信号。

    2. 截尾均值闸：
       丢掉最高 `trim_ratio` 的帧后均值仍 > `tm_cap`，判为不同内容。
       主要用来挡"整体距离都很高"的两个无关视频。

    3. 通过以上两闸后计算分数 = 截尾均值 / (1 + 标准差)：
       同内容不同编码 → 高 std → 分数更低 → 更容易匹配。
    """
    _BASELINE_BITS = 64.0

    p1 = fs1.phash_parts
    p2 = fs2.phash_parts
    if not p1 or not p2:
        return float("inf")
    n = min(len(p1), len(p2))
    if n == 0:
        return float("inf")

    dists: List[float] = []
    for i in range(n):
        try:
            h1 = _hex_to_hash_cached(p1[i])
            h2 = _hex_to_hash_cached(p2[i])
            raw_d = h1 - h2
            bits = h1.hash.size
            norm_d = raw_d * _BASELINE_BITS / max(bits, 1)
            dists.append(norm_d)
        except Exception as e:
            LOGGER.debug("[phash_distance] 计算失败：%s", e)
            return float("inf")

    dists.sort()

    # 1) 双峰差闸
    if n >= 6:
        half = n // 2
        bh_mean = sum(dists[:half]) / half
        th_mean = sum(dists[half:]) / (n - half)
        if (th_mean - bh_mean) > float(bimodal_gap_cap):
            return float("inf")

    # 2) 截尾均值闸
    tr = max(0.0, min(0.5, float(trim_ratio)))
    trim = max(1, int(n * tr))
    trimmed = dists[: n - trim]
    if not trimmed:
        return float("inf")
    tm = sum(trimmed) / len(trimmed)

    if tm > float(tm_cap):
        return float("inf")

    if n < 6:
        return tm

    mean_all = sum(dists) / n
    variance = sum((d - mean_all) ** 2 for d in dists) / (n - 1)
    std = math.sqrt(variance)

    return tm / (1.0 + std)


def cluster_bucket_by_phash(fs_list: List[FileSig], threshold: float,
                            tm_cap: float = 12.0,
                            trim_ratio: float = 0.2,
                            bimodal_gap_cap: float = 40.0,
                            color_distance_threshold: float = 0.15,
                            semantic_distance_threshold: float = 0.015,
                            semantic_max_threshold: float = 0.040,
                            semantic_peak_ratio_threshold: float = 3.8,
                            semantic_peak_min_max: float = 0.015,
                            semantic_drift_p90_exempt: float = 0.005,
                            semantic_drift_sparse_mid_count: int = 2,
                            # —— Patch-level 空间分布闸 ——
                            semantic_patch_enabled: bool = True,
                            semantic_patch_grid: int = 8,
                            semantic_patch_hot_threshold: float = 0.015,
                            semantic_patch_min_hot_patches: int = 12,
                            semantic_patch_center_margin: float = 0.4,
                            semantic_patch_edge_margin: float = 0.6,
                            semantic_patch_corner_merge_frac: float = 0.55,
                            semantic_patch_center_split_frac: float = 0.45,
                            semantic_patch_persistent_frame_frac: float = 0.5,
                            semantic_patch_persistent_min: int = 2,
                            semantic_patch_persistent_max: int = 8,
                            semantic_patch_persistent_corner_min: float = 0.8,
                            semantic_patch_weak_center_max: float = 0.12,
                            semantic_patch_weak_hot_ratio_max: float = 0.10,
                            semantic_patch_heavy_persistent_min: int = 10,
                            semantic_patch_heavy_hot_ratio_min: float = 0.20,
                            semantic_patch_heavy_pers_corner_max: float = 0.85,
                            semantic_patch_heavy_min_ratio: float = 2.5,
                            semantic_patch_center_persistent_corner_max: float = 0.20,
                            semantic_patch_center_persistent_total_corner_max: float = 0.25,
                            semantic_patch_drift_exempt_center_persistent_blocks: bool = True,
                            semantic_patch_anim_hot_ratio_min: float = 0.15,
                            semantic_patch_anim_corner_min: float = 0.65,
                            semantic_patch_anim_center_max: float = 0.10,
                            semantic_patch_anim_max_override_factor: float = 2.5,
                            semantic_patch_ratio_rescue_corner_min: float = 0.60,
                            semantic_patch_ratio_rescue_center_max: float = 0.15,
                            semantic_patch_ratio_rescue_mean_max: float = 0.010,
                            semantic_patch_max_rescue_corner_min: float = 0.75,
                            semantic_patch_max_rescue_center_max: float = 0.10,
                            semantic_patch_max_rescue_mean_max: float = 0.010,
                            semantic_patch_center_mask_enabled: bool = True,
                            semantic_patch_center_mask_inner: int = 4,
                            semantic_patch_center_mask_mean_max: float = 0.006,
                            semantic_patch_center_mask_max_max: float = 0.025,
                            semantic_patch_center_mask_hot_ratio_max: float = 0.06,
                            semantic_patch_center_mask_relaxed_enabled: bool = True,
                            semantic_patch_center_mask_relaxed_hot_ratio_max: float = 0.14,
                            semantic_patch_center_mask_relaxed_dom_q_max: float = 0.30,
                            semantic_patch_center_mask_relaxed_corner_min: float = 0.85,
                            semantic_patch_center_mask_relaxed_center_max: float = 0.05,
                            semantic_patch_center_mask_relaxed_mean_max: float = 0.0065,
                            semantic_patch_center_mask_relaxed_p90_max: float = 0.010,
                            semantic_patch_center_mask_relaxed_max_max: float = 0.060,
                            audio_merge_override_color: bool = False,
                            audio_merge_threshold: float = 0.15,
                            # —— LLM 文本仲裁（灰区双向 override，可选）——
                            llm_enabled: bool = False,
                            llm_handle: Optional[Any] = None,     # 预加载的 LLMHandle；None 则按需加载
                            llm_cache: Optional["SigCache"] = None,
                            llm_model_name: str = "Qwen/Qwen3-8B",
                            llm_model_cache_dir: str = "models_cache",
                            llm_device: str = "auto",
                            llm_quantization: str = "4bit",
                            llm_max_new_tokens: int = 32,
                            llm_max_input_tokens: int = 1024,
                            llm_enable_thinking: bool = False,
                            llm_gray_color_max_excess: float = 0.06,
                            llm_gray_semantic_mean_max_excess: float = 0.008,
                            llm_gray_semantic_max_max_excess: float = 0.02,
                            llm_gray_semantic_peak_max_excess: float = 1.5,
                            llm_intercept_on_patch_uncertain: bool = True,
                            llm_intercept_pixel_clean_mean: float = 0.010,
                            llm_intercept_pixel_clean_max: float = 0.030,
                            llm_rescue_enabled: bool = False,
                            llm_skip_if_no_meta: bool = True) -> List[List[FileSig]]:
    """
    在同一个“粗桶”（时长 + 可选音频）里，对文件做 pHash 模糊聚类：
      - 第一道闸 pHash：组合分 <= threshold 才继续；
      - 第二道闸 颜色直方图：Bhattacharyya 距离 <= color_distance_threshold
        专挡"同构图+同动作，不同角色/服饰" 类误报；
      - 第三道闸 语义 embedding（可选，三个触发条件 OR）：
          * mean(per-frame cosine) > semantic_distance_threshold
              → 整体不同（短视频差分 117/457 类，mean 普遍 > 0.017）
          * max(per-frame cosine) > semantic_max_threshold
              → 存在极端差异帧（绝对值兜底）
          * max/mean > semantic_peak_ratio_threshold 且 max > semantic_peak_min_max
              → 尖峰分布：区分"水印/重编码全局漂移"（平坦, ratio≈2~3.5）
                vs "同源局部差分"（尖峰, ratio≈4~6+）；
                peak_min_max 防止极低 mean 时比率虚高。
        任一条件成立即判为差分，不合并。
        编码漂移例外（两种模式，OR；命中即豁免 max/ratio 尖峰闸，mean 闸仍生效）：
          * 平坦漂移：p90(per-frame) <= semantic_drift_p90_exempt
              → 绝大多数帧几乎完全一致；典型为同源不同码率的 decoder 帧对齐漂移。
          * 稀疏超级尖峰：max > semantic_max_threshold
              且 (0.5×semantic_max_threshold, semantic_max_threshold] 区间帧数
              <= semantic_drift_sparse_mid_count，且 mean <= semantic_distance_threshold
              → 极少数帧飙很高，但中间没有"过渡帧"；典型为水印/关键帧孤立脏帧。
              真差分（某段内容被改）会在中间带留下多帧过渡，中间带帧数普遍 >= 3。
      - 使用并查集得到最终若干子组；只返回大小 >= 2 的子组。
    """
    try:
        from semantic_features import semantic_frame_distances
        semantic_available = True
    except Exception:
        semantic_available = False
        semantic_frame_distances = None  # type: ignore

    try:
        from semantic_features import patch_spatial_verdict  # type: ignore
        patch_available = True
    except Exception:
        patch_available = False
        patch_spatial_verdict = None  # type: ignore

    try:
        from semantic_features import patch_center_masked_distances  # type: ignore
    except Exception:
        patch_center_masked_distances = None  # type: ignore

    # LLM 已从筛重主流程退役：保留参数与兼容代码，避免旧脚本调用报错，
    # 但这里统一熄火，确保整条聚类链路始终是纯像素/纯特征判定。
    llm_enabled = False
    llm_rescue_enabled = False
    llm_intercept_on_patch_uncertain = False

    n = len(fs_list)
    if n < 2:
        return []

    parent = list(range(n))
    rank = [0] * n

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int):
        ra, rb = find(a), find(b)
        if ra == rb:
            return
        if rank[ra] < rank[rb]:
            parent[ra] = rb
        elif rank[ra] > rank[rb]:
            parent[rb] = ra
        else:
            parent[rb] = ra
            rank[ra] += 1

    # —— LLM 灰区仲裁闭包（进程级懒加载 + pair 缓存）——
    # 返回 "same" / "different" / "uncertain" / "skip"：
    #   skip 表示没有元数据或 LLM 不可用，外层应该维持原判决。
    # 优先用调用者传入的 llm_handle；否则退回到模块级单例 _LLM_HANDLE。
    # _LLM_FAILED 一旦置位，本进程余下所有桶都直接跳过 LLM，避免重复失败刷屏。

    def _get_llm() -> Optional[Any]:
        global _LLM_HANDLE, _LLM_FAILED
        if llm_handle is not None:
            return llm_handle
        if _LLM_FAILED:
            return None
        if _LLM_HANDLE is not None:
            return _LLM_HANDLE
        if not llm_enabled:
            return None
        with _LLM_LOCK:
            # 二次检查：可能其他线程在等锁期间已完成加载或标记失败
            if _LLM_FAILED:
                return None
            if _LLM_HANDLE is not None:
                return _LLM_HANDLE
            try:
                from llm_arbitrator import load_llm_model
                _LLM_HANDLE = load_llm_model(
                    llm_model_name,
                    device=llm_device,
                    quantization=llm_quantization,
                    max_new_tokens=int(llm_max_new_tokens),
                    max_input_tokens=int(llm_max_input_tokens),
                    enable_thinking=bool(llm_enable_thinking),
                    cache_dir=llm_model_cache_dir,
                )
                return _LLM_HANDLE
            except Exception as e:
                import traceback as _tb
                LOGGER.error(
                    "[llm] 加载失败：%s；本进程余下桶保持算法原判\n%s",
                    e, _tb.format_exc(),
                )
                _LLM_FAILED = True
                return None

    def _llm_consult(fs_a: FileSig, fs_b: FileSig, context: str) -> str:
        """对一对 FileSig 做文本仲裁，返回 'same' / 'different' / 'uncertain' / 'skip'。"""
        if not llm_enabled:
            return "skip"
        title_a, desc_a = fs_a.title_text, fs_a.description_text
        title_b, desc_b = fs_b.title_text, fs_b.description_text
        if llm_skip_if_no_meta and not (title_a or desc_a) and not (title_b or desc_b):
            return "skip"
        handle = _get_llm()
        if handle is None:
            return "skip"
        try:
            from llm_arbitrator import judge_pair as _judge_pair, pair_hash as _pair_hash
        except Exception:
            return "skip"
        text_hash = _pair_hash(title_a, desc_a, title_b, desc_b)
        wid_a, wid_b = fs_a.item_id, fs_b.item_id
        if llm_cache is not None:
            try:
                cached = llm_cache.llm_get(
                    wid_a, wid_b, handle.model_name, handle.prompt_hash, text_hash,
                )
            except Exception as e:
                LOGGER.debug("[llm] 读缓存失败：%s", e)
                cached = None
            if cached:
                LOGGER.info("[llm:cache %s] %s <-> %s → %s (%s)",
                            context, wid_a, wid_b, cached["verdict"], cached.get("reason", ""))
                return cached["verdict"]
        try:
            t0 = time.time()
            result = _judge_pair(handle, (title_a, desc_a), (title_b, desc_b))
            dt = time.time() - t0
            LOGGER.info("[llm:%.1fs %s] %s <-> %s → %s (%s)",
                        dt, context, wid_a, wid_b,
                        result["verdict"], result.get("reason", ""))
        except Exception as e:
            LOGGER.warning("[llm] 推理异常：%s", e)
            return "skip"
        if llm_cache is not None:
            try:
                llm_cache.llm_upsert(
                    wid_a, wid_b, handle.model_name, handle.prompt_hash, text_hash,
                    result["verdict"], result.get("reason", ""), result.get("raw", ""),
                )
            except Exception as e:
                LOGGER.debug("[llm] 写缓存失败：%s", e)
        return result["verdict"]

    def _llm_rescue_split(i: int, j: int, gate: str) -> bool:
        """SPLIT→MERGE 救回：返回 True 表示 LLM 判 same，外层应该 union。
        llm_rescue_enabled=False 时直接返回 False（不调 LLM，保持像素原判）。
        """
        if not llm_rescue_enabled:
            return False
        v = _llm_consult(fs_list[i], fs_list[j], gate)
        if v == "same":
            LOGGER.info("[llm-rescue:%s] 救回 %s <-> %s（LLM 判 same）",
                        gate, fs_list[i].item_id, fs_list[j].item_id)
            return True
        return False

    def _llm_block_merge(i: int, j: int, gate: str) -> bool:
        """MERGE→SPLIT 拦下：返回 True 表示 LLM 判 different，外层应该 continue。"""
        v = _llm_consult(fs_list[i], fs_list[j], gate)
        if v == "different":
            LOGGER.info("[llm-block:%s] 拦下 %s <-> %s（LLM 判 different）",
                        gate, fs_list[i].item_id, fs_list[j].item_id)
            return True
        return False

    def _llm_patch_confirm_same(i: int, j: int, gate: str) -> bool:
        """patch 触发的候选 rescue：必须 LLM 明确判 same 才放行。
        与 _llm_rescue_split 的区别是此处不受 llm_rescue_enabled 开关控制——
        信号来源是 patch 空间分布提示像水印，是否真是水印由 LLM 文本仲裁决定。
        未启用 LLM / LLM 抽检或判 different 时返回 False，外层退回保守 SPLIT。"""
        if not llm_enabled:
            return False
        v = _llm_consult(fs_list[i], fs_list[j], gate)
        if v == "same":
            LOGGER.info("[llm-patch-confirm:%s] 确认 %s <-> %s（LLM 判 same）",
                        gate, fs_list[i].item_id, fs_list[j].item_id)
            return True
        return False

    def _center_patch_confirm_same(i: int, j: int, gate: str,
                                   patch_hot_ratio_val: float,
                                   patch_corner_frac_val: float,
                                   patch_center_frac_val: float,
                                   patch_dom_q_frac_val: float) -> bool:
        """patch 触发的候选 rescue：在 DINO patch 网格中心 keep×keep 格上复验。
        严格层：中心区域 mean/max 都足够小且全局 hot_ratio 也足够低 → 视为同源。
        二级层：仅给 max-corner-dominant 候选用，允许少量孤立峰值，但要求 center_p90、
        角落/中心形态和 dom_q 都满足更细的约束。否则返回 False 交由上层。

        这是方案 3（中心屏蔽）的纯像素确认；零新签名成本、复用已缓存 patch embedding。
        额外 hot_ratio 判据避免"四角都有不同内容"（G342/G344 型）误救——那类对的
        整体 hot patch 占比 > 0.08，远高于真水印/字幕的 < 0.06。"""
        if (not semantic_patch_center_mask_enabled
                or patch_center_masked_distances is None):
            return False
        pa = fs_list[i].semantic_patch_embeddings
        pb = fs_list[j].semantic_patch_embeddings
        if pa is None or pb is None:
            return False
        stats = patch_center_masked_distances(
            pa, pb,
            grid_side=int(semantic_patch_grid),
            keep_inner=int(semantic_patch_center_mask_inner),
        )
        if stats is None:
            return False
        strict_hot_ok = patch_hot_ratio_val <= float(semantic_patch_center_mask_hot_ratio_max)
        strict_mean_ok = stats["mean"] <= float(semantic_patch_center_mask_mean_max)
        strict_max_ok = stats["max"] <= float(semantic_patch_center_mask_max_max)
        if strict_hot_ok and strict_mean_ok and strict_max_ok:
            LOGGER.info(
                "[center-patch-confirm:%s] 确认 %s <-> %s  "
                "keep=%d hot_r=%.3f domq=%.2f center_mean=%.4f<=%.4f "
                "center_p90=%.4f center_max=%.4f<=%.4f",
                gate, fs_list[i].item_id, fs_list[j].item_id,
                int(semantic_patch_center_mask_inner),
                patch_hot_ratio_val,
                patch_dom_q_frac_val,
                stats["mean"], float(semantic_patch_center_mask_mean_max),
                stats["p90"],
                stats["max"], float(semantic_patch_center_mask_max_max),
            )
            return True
        relaxed_gate = (
            bool(semantic_patch_center_mask_relaxed_enabled)
            and gate == "max-corner-dominant"
            and patch_hot_ratio_val <= float(semantic_patch_center_mask_relaxed_hot_ratio_max)
            and patch_corner_frac_val >= float(semantic_patch_center_mask_relaxed_corner_min)
            and patch_center_frac_val <= float(semantic_patch_center_mask_relaxed_center_max)
            and patch_dom_q_frac_val <= float(semantic_patch_center_mask_relaxed_dom_q_max)
        )
        relaxed_mean_ok = stats["mean"] <= float(semantic_patch_center_mask_relaxed_mean_max)
        relaxed_p90_ok = stats["p90"] <= float(semantic_patch_center_mask_relaxed_p90_max)
        relaxed_max_ok = stats["max"] <= float(semantic_patch_center_mask_relaxed_max_max)
        if relaxed_gate and relaxed_mean_ok and relaxed_p90_ok and relaxed_max_ok:
            LOGGER.info(
                "[center-patch-confirm:%s/relaxed] 确认 %s <-> %s  "
                "keep=%d hot_r=%.3f<=%.3f domq=%.2f<=%.2f corner=%.2f>=%.2f "
                "center=%.2f<=%.2f center_mean=%.4f<=%.4f center_p90=%.4f<=%.4f "
                "center_max=%.4f<=%.4f",
                gate, fs_list[i].item_id, fs_list[j].item_id,
                int(semantic_patch_center_mask_inner),
                patch_hot_ratio_val, float(semantic_patch_center_mask_relaxed_hot_ratio_max),
                patch_dom_q_frac_val, float(semantic_patch_center_mask_relaxed_dom_q_max),
                patch_corner_frac_val, float(semantic_patch_center_mask_relaxed_corner_min),
                patch_center_frac_val, float(semantic_patch_center_mask_relaxed_center_max),
                stats["mean"], float(semantic_patch_center_mask_relaxed_mean_max),
                stats["p90"], float(semantic_patch_center_mask_relaxed_p90_max),
                stats["max"], float(semantic_patch_center_mask_relaxed_max_max),
            )
            return True
        LOGGER.info(
            "[center-patch-confirm:%s/reject] %s <-> %s  keep=%d hot_r=%.3f(<=%.3f? %s) "
            "domq=%.2f corner=%.2f center=%.2f center_mean=%.4f(<=%.4f? %s) "
            "center_p90=%.4f center_max=%.4f(<=%.4f? %s) relaxed=%s "
            "[relaxed hot<=%.3f corner>=%.2f center<=%.2f domq<=%.2f mean<=%.4f p90<=%.4f max<=%.4f]",
            gate, fs_list[i].item_id, fs_list[j].item_id,
            int(semantic_patch_center_mask_inner),
            patch_hot_ratio_val, float(semantic_patch_center_mask_hot_ratio_max), strict_hot_ok,
            patch_dom_q_frac_val, patch_corner_frac_val, patch_center_frac_val,
            stats["mean"], float(semantic_patch_center_mask_mean_max), strict_mean_ok,
            stats["p90"],
            stats["max"], float(semantic_patch_center_mask_max_max), strict_max_ok,
            relaxed_gate,
            float(semantic_patch_center_mask_relaxed_hot_ratio_max),
            float(semantic_patch_center_mask_relaxed_corner_min),
            float(semantic_patch_center_mask_relaxed_center_max),
            float(semantic_patch_center_mask_relaxed_dom_q_max),
            float(semantic_patch_center_mask_relaxed_mean_max),
            float(semantic_patch_center_mask_relaxed_p90_max),
            float(semantic_patch_center_mask_relaxed_max_max),
        )
        return False

    def _patch_rescue_confirm_same(i: int, j: int, gate: str,
                                   patch_hot_ratio_val: float,
                                   patch_corner_frac_val: float,
                                   patch_center_frac_val: float,
                                   patch_dom_q_frac_val: float) -> bool:
        """patch 触发的候选 rescue 统一入口：仅使用中心 patch mask 复验。"""
        if _center_patch_confirm_same(
            i, j, gate,
            patch_hot_ratio_val,
            patch_corner_frac_val,
            patch_center_frac_val,
            patch_dom_q_frac_val,
        ):
            return True
        return False

    # 两两比较并 union
    for i in range(n):
        for j in range(i + 1, n):
            d = phash_distance(
                fs_list[i], fs_list[j],
                tm_cap=tm_cap, trim_ratio=trim_ratio,
                bimodal_gap_cap=bimodal_gap_cap,
            )
            if d > threshold:
                continue
            # 颜色二次闸：两侧均有颜色直方图时才启用，缺数据就退化为仅 pHash 判定
            ch1 = fs_list[i].color_histogram
            ch2 = fs_list[j].color_histogram
            if ch1 and ch2:
                cd = color_histogram_distance(ch1, ch2)
                if not math.isnan(cd) and cd > color_distance_threshold:
                    # 音频 rescue：同一视频加滤镜/调色重编码会让 color 超阈值，但 Chromaprint 几乎不变。
                    # 仅当开关开启、两侧均有原始 fp 且距离在阈值内时，豁免 color 的拆分决定。
                    audio_rescue = False
                    if audio_merge_override_color:
                        ad = chromaprint_distance(
                            fs_list[i].audio_fp_raw, fs_list[j].audio_fp_raw,
                        )
                        if ad is not None and ad <= audio_merge_threshold:
                            audio_rescue = True
                            LOGGER.info(
                                "[audio-rescue] 救回 %s <-> %s  color=%.3f > %.3f  audio_dist=%.4f <= %.4f",
                                fs_list[i].item_id, fs_list[j].item_id,
                                cd, color_distance_threshold, ad, audio_merge_threshold,
                            )
                        elif ad is not None:
                            LOGGER.info(
                                "[audio-rescue] 未救 %s <-> %s  color=%.3f audio_dist=%.4f > %.4f",
                                fs_list[i].item_id, fs_list[j].item_id,
                                cd, ad, audio_merge_threshold,
                            )
                    if not audio_rescue:
                        # LLM 灰区救回（SPLIT→MERGE）：色差刚越线且有元数据时问 LLM
                        if (llm_enabled
                                and (cd - color_distance_threshold) <= float(llm_gray_color_max_excess)
                                and _llm_rescue_split(i, j, "color-gate")):
                            union(i, j)
                            continue
                        LOGGER.info(
                            "[color-gate] 挡下 %s <-> %s  phash=%.3f color_dist=%.3f > %.3f",
                            fs_list[i].item_id, fs_list[j].item_id, d, cd, color_distance_threshold,
                        )
                        continue
            # 语义三次闸（mean / max / peak-ratio 三元 OR）：
            #   - mean  闸：整体性差异（117/457 类短视频局部差分，mean 普遍 > 0.017）
            #   - max   闸：存在明显偏离帧（绝对值兜底）
            #   - ratio 闸：max/mean 衡量分布尖峰度——区分水印（平坦 ~2~3.5）和局部差分（尖峰 >3.9）
            emb1 = fs_list[i].semantic_embeddings
            emb2 = fs_list[j].semantic_embeddings
            if (semantic_available
                    and emb1 is not None and emb2 is not None
                    and getattr(emb1, "size", 0) > 0 and getattr(emb2, "size", 0) > 0):
                frame_d = semantic_frame_distances(emb1, emb2)
                if frame_d is not None and frame_d.size > 0:
                    sd_mean = float(frame_d.mean())
                    sd_max = float(frame_d.max())
                    sd_p90 = float(np.percentile(frame_d, 90))
                    sd_ratio = sd_max / max(sd_mean, 1e-9)
                    # 编码漂移例外（平坦漂移 OR 稀疏超级尖峰）：
                    #  - 平坦漂移：p90 极低说明 90% 帧几乎完全一致
                    #  - 稀疏超级尖峰：max 很高但"过渡带"帧数很少，mean 也小
                    #    → 极少数孤立脏帧（水印突变/关键帧漂移），不是真差分
                    mid_lower = semantic_max_threshold * 0.5
                    mid_band_count = int(
                        np.logical_and(frame_d > mid_lower,
                                       frame_d <= semantic_max_threshold).sum()
                    )
                    n_above_max = int((frame_d > semantic_max_threshold).sum())
                    sparse_peak_exempt = (
                        sd_max > semantic_max_threshold
                        and mid_band_count <= semantic_drift_sparse_mid_count
                        and n_above_max <= semantic_drift_sparse_mid_count
                        and sd_mean <= semantic_distance_threshold
                    )
                    drift_exempt = (sd_p90 <= semantic_drift_p90_exempt) or sparse_peak_exempt
                    if sd_mean > semantic_distance_threshold:
                        # LLM 灰区救回：mean 刚越线才问，远离阈值不浪费 token
                        if (llm_enabled
                                and (sd_mean - semantic_distance_threshold)
                                    <= float(llm_gray_semantic_mean_max_excess)
                                and _llm_rescue_split(i, j, "sem-mean")):
                            union(i, j)
                            continue
                        LOGGER.info(
                            "[semantic-gate/mean]  挡下 %s <-> %s  phash=%.3f mean=%.4f > %.4f  (max=%.4f p90=%.4f r=%.2f)",
                            fs_list[i].item_id, fs_list[j].item_id, d,
                            sd_mean, semantic_distance_threshold, sd_max, sd_p90, sd_ratio,
                        )
                        continue
                    if drift_exempt:
                        # drift_exempt 本意是"编码漂移型"真合并对，但如果 patch 层识别出
                        # 持久的中心差异（content_diff_center_persistent），说明中心
                        # 确实存在稳定内容变化（S43 型），仍应挡下。只运行 patch 评估
                        # 不做任何其它判决覆盖。
                        if (semantic_patch_drift_exempt_center_persistent_blocks
                                and semantic_patch_enabled and patch_available
                                and fs_list[i].semantic_patch_embeddings is not None
                                and fs_list[j].semantic_patch_embeddings is not None):
                            _pv = patch_spatial_verdict(
                                fs_list[i].semantic_patch_embeddings,
                                fs_list[j].semantic_patch_embeddings,
                                grid_side=int(semantic_patch_grid),
                                hot_threshold=float(semantic_patch_hot_threshold),
                                min_hot_patches=int(semantic_patch_min_hot_patches),
                                center_margin=float(semantic_patch_center_margin),
                                edge_margin=float(semantic_patch_edge_margin),
                                corner_merge_frac=float(semantic_patch_corner_merge_frac),
                                center_split_frac=float(semantic_patch_center_split_frac),
                                persistent_frame_frac=float(semantic_patch_persistent_frame_frac),
                                persistent_min=int(semantic_patch_persistent_min),
                                persistent_max=int(semantic_patch_persistent_max),
                                persistent_corner_min=float(semantic_patch_persistent_corner_min),
                                weak_center_max=float(semantic_patch_weak_center_max),
                                weak_hot_ratio_max=float(semantic_patch_weak_hot_ratio_max),
                                heavy_persistent_min=int(semantic_patch_heavy_persistent_min),
                                heavy_hot_ratio_min=float(semantic_patch_heavy_hot_ratio_min),
                                heavy_pers_corner_max=float(semantic_patch_heavy_pers_corner_max),
                                center_persistent_corner_max=float(semantic_patch_center_persistent_corner_max),
                                center_persistent_total_corner_max=float(semantic_patch_center_persistent_total_corner_max),
                                anim_hot_ratio_min=float(semantic_patch_anim_hot_ratio_min),
                                anim_corner_min=float(semantic_patch_anim_corner_min),
                                anim_center_max=float(semantic_patch_anim_center_max),
                            )
                            if _pv["kind"] == "content_diff_center_persistent":
                                LOGGER.info(
                                    "[patch-gate/drift-exempt→center_persistent] 挡下 %s <-> %s  "
                                    "phash=%.3f  (p90=%.4f max=%.4f mean=%.4f | kind=%s hot=%d "
                                    "hot_r=%.3f corner=%.2f center=%.2f pers=%d pc=%.2f)",
                                    fs_list[i].item_id, fs_list[j].item_id, d,
                                    sd_p90, sd_max, sd_mean, _pv["kind"], _pv["hot"],
                                    _pv["hot_ratio"], _pv["corner_frac"], _pv["center_frac"],
                                    _pv["persistent"], _pv["pers_corner_frac"],
                                )
                                continue
                        if sd_max > semantic_max_threshold or (
                            sd_ratio > semantic_peak_ratio_threshold and sd_max > semantic_peak_min_max
                        ):
                            kind = "flat(p90)" if sd_p90 <= semantic_drift_p90_exempt else "sparse-peak"
                            LOGGER.info(
                                "[semantic-gate/drift-exempt:%s] 放行 %s <-> %s  phash=%.3f "
                                "p90=%.4f max=%.4f mid(%.3f,%.3f]=%d n>max=%d mean=%.4f r=%.2f",
                                kind, fs_list[i].item_id, fs_list[j].item_id, d,
                                sd_p90, sd_max, mid_lower, semantic_max_threshold,
                                mid_band_count, n_above_max, sd_mean, sd_ratio,
                            )
                    else:
                        # —— Patch 仲裁：mean 通过后总是运行（可用时），作为第四道独立闸 ——
                        # 五种决策：
                        #   watermark_strong：可推翻 max 闸 + ratio 闸（真固定水印）
                        #   watermark_weak：仅推翻 ratio 闸（低码率重编码边缘抖动）
                        #   content_diff_heavy：即便语义三闸都未触发也 SPLIT（S44 型大面积稳定差异）
                        #   content_diff_center_persistent：持久热点在中心 → SPLIT（S43 型微弱中心差异）
                        #   其他（uncertain / content_diff / disabled）：保守维持语义闸判决
                        max_gate_hit = sd_max > semantic_max_threshold
                        ratio_gate_hit = (sd_ratio > semantic_peak_ratio_threshold
                                          and sd_max > semantic_peak_min_max)
                        patch_verdict_kind = "disabled"
                        patch_stats_str = "disabled"
                        patch_corner_frac = 0.0
                        patch_center_frac = 0.0
                        patch_hot_ratio = 0.0
                        patch_dom_q_frac = 0.0
                        if (semantic_patch_enabled and patch_available
                                and fs_list[i].semantic_patch_embeddings is not None
                                and fs_list[j].semantic_patch_embeddings is not None):
                            pv = patch_spatial_verdict(
                                fs_list[i].semantic_patch_embeddings,
                                fs_list[j].semantic_patch_embeddings,
                                grid_side=int(semantic_patch_grid),
                                hot_threshold=float(semantic_patch_hot_threshold),
                                min_hot_patches=int(semantic_patch_min_hot_patches),
                                center_margin=float(semantic_patch_center_margin),
                                edge_margin=float(semantic_patch_edge_margin),
                                corner_merge_frac=float(semantic_patch_corner_merge_frac),
                                center_split_frac=float(semantic_patch_center_split_frac),
                                persistent_frame_frac=float(semantic_patch_persistent_frame_frac),
                                persistent_min=int(semantic_patch_persistent_min),
                                persistent_max=int(semantic_patch_persistent_max),
                                persistent_corner_min=float(semantic_patch_persistent_corner_min),
                                weak_center_max=float(semantic_patch_weak_center_max),
                                weak_hot_ratio_max=float(semantic_patch_weak_hot_ratio_max),
                                heavy_persistent_min=int(semantic_patch_heavy_persistent_min),
                                heavy_hot_ratio_min=float(semantic_patch_heavy_hot_ratio_min),
                                heavy_pers_corner_max=float(semantic_patch_heavy_pers_corner_max),
                                center_persistent_corner_max=float(semantic_patch_center_persistent_corner_max),
                                center_persistent_total_corner_max=float(semantic_patch_center_persistent_total_corner_max),
                                anim_hot_ratio_min=float(semantic_patch_anim_hot_ratio_min),
                                anim_corner_min=float(semantic_patch_anim_corner_min),
                                anim_center_max=float(semantic_patch_anim_center_max),
                            )
                            patch_verdict_kind = pv["kind"]
                            patch_corner_frac = float(pv["corner_frac"])
                            patch_center_frac = float(pv["center_frac"])
                            patch_hot_ratio = float(pv["hot_ratio"])
                            patch_dom_q_frac = float(pv.get("dom_q_frac", 0.0))
                            patch_stats_str = (
                                f"kind={patch_verdict_kind} hot={pv['hot']} "
                                f"hot_r={pv['hot_ratio']:.3f} corner={pv['corner_frac']:.2f} "
                                f"center={pv['center_frac']:.2f} pers={pv['persistent']} "
                                f"pc={pv['pers_corner_frac']:.2f} domq={pv.get('dom_q_frac', 0.0):.2f} "
                                f"grid={pv['grid']}"
                            )

                        if max_gate_hit:
                            if patch_verdict_kind == "watermark_strong":
                                LOGGER.info(
                                    "[semantic-gate/max→patch:strong-wm] 推翻 max 闸放行 %s <-> %s  "
                                    "phash=%.3f max=%.4f mean=%.4f r=%.2f | %s",
                                    fs_list[i].item_id, fs_list[j].item_id, d,
                                    sd_max, sd_mean, sd_ratio, patch_stats_str,
                                )
                                union(i, j)
                                continue
                            # watermark_anim：动画水印同样可以推翻 max 闸，但需上限保护
                            # （防止真大差异被误救）。max 不得超过阈值 × override_factor。
                            # 典型 anim 水印对的 sd_max 在 0.04~0.09 之间；PERS=0 说明热点在角落
                            # 逐帧漂移，与真差分（持久中心热点）本质不同。
                            if (patch_verdict_kind == "watermark_anim"
                                    and sd_max <= semantic_max_threshold
                                        * float(semantic_patch_anim_max_override_factor)):
                                LOGGER.info(
                                    "[semantic-gate/max→patch:anim-wm] 推翻 max 闸放行 %s <-> %s  "
                                    "phash=%.3f max=%.4f mean=%.4f r=%.2f | %s",
                                    fs_list[i].item_id, fs_list[j].item_id, d,
                                    sd_max, sd_mean, sd_ratio, patch_stats_str,
                                )
                                union(i, j)
                                continue
                            # patch=uncertain 但空间分布呈极度角落主导（corner 比 ratio rescue
                            # 的阈值更严苛）+ 中心极干净 + mean 很低 + max 未超过 anim 上限 →
                            # 视为 de facto 动画水印候选。典型场景：1080p vs 4K
                            # 重编码引入的角落压缩噪声，hot_ratio 卡在 0.13 附近没能进 anim 分类。
                            # 注意：像素层分不清这是"真水印"还是"四角都有不同内容"
                            # （G342/G344 型），因此必须走复验：首选中心 patch mask
                            # （对 DINO patch grid 屏蔽外圈两层，复查内圈 mean/max）；
                            # 中心不通过且 LLM 启用时退回 LLM 文本仲裁；两者都不过则保守 SPLIT。
                            if (patch_verdict_kind == "uncertain"
                                    and patch_corner_frac >= float(semantic_patch_max_rescue_corner_min)
                                    and patch_center_frac <= float(semantic_patch_max_rescue_center_max)
                                    and sd_mean <= float(semantic_patch_max_rescue_mean_max)
                                    and sd_max <= semantic_max_threshold
                                        * float(semantic_patch_anim_max_override_factor)
                                    and _patch_rescue_confirm_same(
                                        i, j, "max-corner-dominant",
                                        patch_hot_ratio,
                                        patch_corner_frac,
                                        patch_center_frac,
                                        patch_dom_q_frac)):
                                LOGGER.info(
                                    "[semantic-gate/max→patch:uncertain-corner-dominant] "
                                    "推翻 max 闸放行 %s <-> %s  phash=%.3f max=%.4f mean=%.4f "
                                    "r=%.2f | %s",
                                    fs_list[i].item_id, fs_list[j].item_id, d,
                                    sd_max, sd_mean, sd_ratio, patch_stats_str,
                                )
                                union(i, j)
                                continue
                            if (llm_enabled
                                    and (sd_max - semantic_max_threshold)
                                        <= float(llm_gray_semantic_max_max_excess)
                                    and _llm_rescue_split(i, j, "sem-max")):
                                union(i, j)
                                continue
                            LOGGER.info(
                                "[semantic-gate/max]   挡下 %s <-> %s  phash=%.3f max=%.4f > %.4f  "
                                "(mean=%.4f p90=%.4f r=%.2f | patch=%s)",
                                fs_list[i].item_id, fs_list[j].item_id, d,
                                sd_max, semantic_max_threshold, sd_mean, sd_p90, sd_ratio,
                                patch_stats_str,
                            )
                            continue
                        if ratio_gate_hit:
                            if patch_verdict_kind in (
                                    "watermark_strong", "watermark_weak", "watermark_anim"):
                                LOGGER.info(
                                    "[semantic-gate/peak→patch:%s] 推翻 ratio 闸放行 %s <-> %s  "
                                    "phash=%.3f ratio=%.2f max=%.4f mean=%.4f | %s",
                                    patch_verdict_kind, fs_list[i].item_id, fs_list[j].item_id, d,
                                    sd_ratio, sd_max, sd_mean, patch_stats_str,
                                )
                                union(i, j)
                                continue
                            # patch=uncertain 但空间分布像水印（角落主导 + 中心干净 + mean 很低）
                            # → 视为 de facto 水印候选。patch_spatial_verdict 没把它判成
                            # watermark_weak 通常是因为 hot_ratio 或 center 恰好卡在阈值边缘
                            # （比如 hot_r=0.12 略过 weak_hot_ratio_max=0.10；center=0.12 恰等于
                            # weak_center_max）；本路径只对 ratio 闸放行，不碰 max 闸。
                            # 注意：与 G342/G344 型"四角都有不同内容"（domq≈0.3~0.45，
                            # corner 也高 center 也低 mean 也低）在纯像素特征上无法严格区分，
                            # 所以必须走复验：首选中心 patch mask（屏蔽外圈两层查内圈语义），
                            # 不通过且 LLM 启用时退回 LLM 文本仲裁；两者都不过则保守 SPLIT。
                            if (patch_verdict_kind == "uncertain"
                                    and patch_corner_frac >= float(semantic_patch_ratio_rescue_corner_min)
                                    and patch_center_frac <= float(semantic_patch_ratio_rescue_center_max)
                                    and sd_mean <= float(semantic_patch_ratio_rescue_mean_max)
                                    and _patch_rescue_confirm_same(
                                        i, j, "ratio-corner-dominant",
                                        patch_hot_ratio,
                                        patch_corner_frac,
                                        patch_center_frac,
                                        patch_dom_q_frac)):
                                LOGGER.info(
                                    "[semantic-gate/peak→patch:uncertain-corner-dominant] "
                                    "推翻 ratio 闸放行 %s <-> %s  phash=%.3f ratio=%.2f max=%.4f "
                                    "mean=%.4f | %s",
                                    fs_list[i].item_id, fs_list[j].item_id, d,
                                    sd_ratio, sd_max, sd_mean, patch_stats_str,
                                )
                                union(i, j)
                                continue
                            if (llm_enabled
                                    and (sd_ratio - semantic_peak_ratio_threshold)
                                        <= float(llm_gray_semantic_peak_max_excess)
                                    and _llm_rescue_split(i, j, "sem-peak")):
                                union(i, j)
                                continue
                            LOGGER.info(
                                "[semantic-gate/peak]  挡下 %s <-> %s  phash=%.3f max/mean=%.2f > %.2f  "
                                "(max=%.4f p90=%.4f mean=%.4f | patch=%s)",
                                fs_list[i].item_id, fs_list[j].item_id, d,
                                sd_ratio, semantic_peak_ratio_threshold, sd_max, sd_p90, sd_mean,
                                patch_stats_str,
                            )
                            continue
                        # 语义三闸全通过，但 patch 仍可否决（content_diff_heavy / center_persistent）
                        if patch_verdict_kind in ("content_diff_heavy",
                                                  "content_diff_center_persistent"):
                            # heavy 需要 ratio 足够尖峰（防止 P20 型"patch 多但分布平坦"的重编码水印被误拆）
                            if (patch_verdict_kind == "content_diff_heavy"
                                    and sd_ratio < float(semantic_patch_heavy_min_ratio)):
                                LOGGER.info(
                                    "[patch-gate/heavy→降级] 保留合并 %s <-> %s  phash=%.3f "
                                    "ratio=%.2f < %.2f（分布过平坦，视为重编码漂移）| %s",
                                    fs_list[i].item_id, fs_list[j].item_id, d,
                                    sd_ratio, float(semantic_patch_heavy_min_ratio),
                                    patch_stats_str,
                                )
                                # 降级合并：真在灰区——LLM 有拦下机会
                                if (llm_enabled and llm_intercept_on_patch_uncertain
                                        and _llm_block_merge(i, j, "patch-heavy-degraded")):
                                    continue
                                # 落入 union
                            else:
                                LOGGER.info(
                                    "[patch-gate/content-diff:%s] 挡下 %s <-> %s  phash=%.3f  "
                                    "(max=%.4f mean=%.4f r=%.2f | %s)",
                                    patch_verdict_kind, fs_list[i].item_id, fs_list[j].item_id, d,
                                    sd_max, sd_mean, sd_ratio, patch_stats_str,
                                )
                                continue
                        # 语义+patch 全过但 patch 非明确干净/水印 → 最后一次灰区拦下机会
                        elif (llm_enabled and llm_intercept_on_patch_uncertain
                                and patch_verdict_kind in (
                                    "content_diff", "uncertain", "watermark_weak")):
                            # pixel-clean 短路：像素两边信号都很低时就不再问 LLM。
                            # 动机：LLM 只看标题/描述，遇到同视频不同命名习惯（汉化/译名/重打包）
                            # 往往误判 different；这种情况下让 LLM 拥有否决权反而是噪声。
                            pixel_clean = (
                                sd_mean <= float(llm_intercept_pixel_clean_mean)
                                and sd_max <= float(llm_intercept_pixel_clean_max)
                            )
                            if pixel_clean:
                                LOGGER.info(
                                    "[llm-intercept/pixel-clean] 放行 %s <-> %s  phash=%.3f "
                                    "mean=%.4f<=%.4f max=%.4f<=%.4f | patch=%s",
                                    fs_list[i].item_id, fs_list[j].item_id, d,
                                    sd_mean, float(llm_intercept_pixel_clean_mean),
                                    sd_max, float(llm_intercept_pixel_clean_max),
                                    patch_stats_str,
                                )
                            elif _llm_block_merge(i, j, f"patch-{patch_verdict_kind}"):
                                continue
            union(i, j)

    clusters_dict: Dict[int, List[FileSig]] = defaultdict(list)
    for idx in range(n):
        root = find(idx)
        clusters_dict[root].append(fs_list[idx])

    # 只保留至少 2 个文件的子组
    return [lst for lst in clusters_dict.values() if len(lst) >= 2]


# ----------------------------- 人工覆写 -----------------------------

def _parse_pair_list(raw: List[str]) -> List[Tuple[str, str]]:
    """把 "wid_a|wid_b" / "wid_a,wid_b" 形式的字符串列表解析成 (a, b) 列表。
    忽略格式不对/空行；顺序保留（a<b 归一化以便去重）。
    """
    out: List[Tuple[str, str]] = []
    seen: Set[Tuple[str, str]] = set()
    for s in raw or []:
        parts = [x.strip() for x in str(s).replace(",", "|").split("|") if x.strip()]
        if len(parts) < 2:
            continue
        a, b = parts[0], parts[1]
        if a == b:
            continue
        key = (a, b) if a < b else (b, a)
        if key in seen:
            continue
        seen.add(key)
        out.append(key)
    return out


def _apply_force_overrides(groups: List[List[FileSig]],
                           force_merge_raw: List[str],
                           force_split_raw: List[str],
                           all_file_sigs_by_item: Dict[str, List[FileSig]]
                           ) -> List[List[FileSig]]:
    """对 cluster 列表做人工覆写。流程：先 split（剔除冲突对），再 merge（union）。
    * force_split：若某 cluster 同时含 A 和 B 两个 item，把含 B 的 FileSig 抽成独立 cluster。
    * force_merge：若 A 与 B 在不同 cluster/未在任何 cluster，union 到一起；单元素也会被纳入。
    只对"wid 在分桶阶段能拿到 FileSig"的对生效；不在桶里的 wid（时长缺失等）会被跳过并记录日志。
    """
    split_pairs = _parse_pair_list(force_split_raw)
    merge_pairs = _parse_pair_list(force_merge_raw)
    if not split_pairs and not merge_pairs:
        return groups

    # Phase 1: force-split —— 遍历所有 cluster，对每个命中 split 对就把 B 的 FileSig 抽出
    out_groups: List[List[FileSig]] = []
    for g in groups:
        items_here = {fs.item_id for fs in g}
        hits = [(a, b) for (a, b) in split_pairs if a in items_here and b in items_here]
        if not hits:
            out_groups.append(g); continue
        # 把被标记的 b 端全部剔出（多对 split 依次剔）
        evicted_items: Set[str] = set()
        for (a, b) in hits:
            evicted_items.add(b)
            LOGGER.info("[force-split] 从 cluster 抽出 %s（保留 %s）", b, a)
        keep = [fs for fs in g if fs.item_id not in evicted_items]
        if len(keep) >= 2:
            out_groups.append(keep)
        elif keep:
            out_groups.append(keep)
        for iid in evicted_items:
            evicted = [fs for fs in g if fs.item_id == iid]
            if evicted:
                out_groups.append(evicted)

    # Phase 2: force-merge —— union-find 合并包含 A/B 的 cluster；允许拉入未在 cluster 的单 item
    idx_by_item: Dict[str, Set[int]] = defaultdict(set)
    for i, g in enumerate(out_groups):
        for fs in g:
            idx_by_item[fs.item_id].add(i)

    parent = list(range(len(out_groups)))
    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(u: int, v: int) -> None:
        ru, rv = find(u), find(v)
        if ru != rv:
            parent[ru] = rv

    def _ensure_cluster_for(iid: str) -> Optional[int]:
        """若 iid 不在任何 cluster 里，从 all_file_sigs_by_item 取 FileSig 新建独立 cluster；
        找不到 FileSig 返回 None。"""
        if iid in idx_by_item and idx_by_item[iid]:
            return next(iter(idx_by_item[iid]))
        fs_list = all_file_sigs_by_item.get(iid) or []
        if not fs_list:
            return None
        new_idx = len(out_groups)
        out_groups.append(list(fs_list))
        parent.append(new_idx)
        idx_by_item[iid].add(new_idx)
        return new_idx

    for (a, b) in merge_pairs:
        ia = _ensure_cluster_for(a)
        ib = _ensure_cluster_for(b)
        if ia is None and ib is None:
            LOGGER.warning("[force-merge] 跳过 %s + %s：双方都不在签名集中", a, b); continue
        if ia is None or ib is None:
            miss = a if ia is None else b
            LOGGER.warning("[force-merge] 跳过 %s + %s：%s 不在签名集中", a, b, miss); continue
        if find(ia) != find(ib):
            union(ia, ib)
            LOGGER.info("[force-merge] union %s + %s", a, b)

    # 收敛为最终 cluster 列表
    collapsed: Dict[int, List[FileSig]] = {}
    for i, g in enumerate(out_groups):
        r = find(i)
        collapsed.setdefault(r, []).extend(g)
    # 至少 2 个不同 item 才成为一组（单 item 多文件不算重复）
    result: List[List[FileSig]] = []
    for lst in collapsed.values():
        unique_items = {fs.item_id for fs in lst}
        if len(unique_items) >= 2:
            result.append(lst)
    return result


# ----------------------------- 重叠 cluster 合并 -----------------------------

def _merge_overlapping_clusters(groups: List[List[FileSig]]) -> List[List[FileSig]]:
    """按 item_id 把有交集的 cluster 合并成一个。用于跨桶交叉比较产生的重叠 cluster 去重。

    实现：把每个 cluster 视为并查集的一个节点，任意 item_id 同时出现在两个 cluster 里
    就 union 它们；最后按 root 归并、并在归并时按 (item_id, path_str) 去重。
    只保留 size >= 2 的子组（与 cluster_bucket_by_phash 的语义保持一致）。
    """
    if not groups:
        return []
    n = len(groups)
    parent = list(range(n))

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    item_to_idx: Dict[str, int] = {}
    for idx, grp in enumerate(groups):
        for fs in grp:
            prev = item_to_idx.get(fs.item_id)
            if prev is None:
                item_to_idx[fs.item_id] = idx
            else:
                union(idx, prev)

    merged: Dict[int, List[FileSig]] = defaultdict(list)
    seen: Dict[int, set] = defaultdict(set)
    for idx, grp in enumerate(groups):
        root = find(idx)
        for fs in grp:
            key = (fs.item_id, str(fs.path))
            if key in seen[root]:
                continue
            seen[root].add(key)
            merged[root].append(fs)

    return [g for g in merged.values() if len(g) >= 2]


# ----------------------------- 导出 -----------------------------

def export_xlsx(groups: List[List[str]], out_xlsx: Path,
                sheet_title: str = "duplicates",
                col_prefix: str = "链接",
                hyperlink: bool = True):
    """
    将每组一行写成 xlsx。
      - hyperlink=True 时每格写超链接并把字体染成蓝色下划线（URL 版用）。
        对本地文件夹路径也有效——Excel 点击会在资源管理器中打开。
      - hyperlink=False 退化成纯文本（当 Excel 对未知协议报警时可以关掉）。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    max_cols = max((len(g) for g in groups), default=0)
    ws.cell(row=1, column=1, value="序号")
    for c in range(2, 2 + max_cols):
        ws.cell(row=1, column=c, value=f"{col_prefix}{c-1}")
    for r, g in enumerate(groups, start=2):
        ws.cell(row=r, column=1, value=(r - 1))
        for j, val in enumerate(g, start=2):
            cell = ws.cell(row=r, column=j, value=val)
            if hyperlink and val:
                try:
                    cell.hyperlink = val
                    cell.font = Font(color="0000FF", underline="single")
                except Exception:
                    pass
    wb.save(out_xlsx)

# ----------------------------- 主流程 -----------------------------

def load_config(path: Optional[Path]) -> Config:
    data = {}
    if path and path.exists():
        with path.open("rb") as f:
            data = tomllib.load(f)
    def get(k, default): return data.get(k, default)
    return Config(
        workshop_root = get("workshop_root", ""),
        output_dir    = get("output_dir", "output"),
        model_cache_dir = str(get("model_cache_dir", "models_cache")),
        we_install_dir = str(get("we_install_dir", "") or ""),
        include_myprojects = bool(get("include_myprojects", True)),
        ffmpeg_path   = get("ffmpeg_path", "ffmpeg"),
        ffprobe_path  = get("ffprobe_path", "ffprobe"),
        fpcalc_path   = get("fpcalc_path", "fpcalc"),

        sample_frames = int(get("sample_frames", 36)),
        phash_size    = int(get("phash_size", 12)),
        audio_window_seconds = int(get("audio_window_seconds", 60)),
        video_window_seconds = int(get("video_window_seconds", 15)),
        seek_ratio = float(get("seek_ratio", 0.5)),

        duration_rounding = get("duration_rounding", "nearest_0.5"),
        require_both_signatures = bool(get("require_both_signatures", True)),
        duration_cross_bucket_tolerance = float(get("duration_cross_bucket_tolerance", 0.6)),

        phash_distance_threshold = float(get("phash_distance_threshold", 1.5)),
        phash_trimmed_mean_cap = float(get("phash_trimmed_mean_cap", 12.0)),
        phash_trim_ratio = float(get("phash_trim_ratio", 0.2)),
        phash_bimodal_gap_cap = float(get("phash_bimodal_gap_cap", 40.0)),

        color_hist_bins_h = int(get("color_hist_bins_h", 16)),
        color_hist_bins_s = int(get("color_hist_bins_s", 4)),
        color_distance_threshold = float(get("color_distance_threshold", 0.15)),

        audio_merge_override_color = bool(get("audio_merge_override_color", False)),
        audio_merge_threshold = float(get("audio_merge_threshold", 0.15)),

        semantic_feature_enabled = bool(get("semantic_feature_enabled", False)),
        semantic_feature_model = str(get("semantic_feature_model", "dinov2_s")),
        semantic_feature_device = str(get("semantic_feature_device", "auto")),
        semantic_sample_frames = int(get("semantic_sample_frames", 60)),
        semantic_distance_threshold = float(get("semantic_distance_threshold", 0.015)),
        semantic_max_threshold = float(get("semantic_max_threshold", 0.040)),
        semantic_peak_ratio_threshold = float(get("semantic_peak_ratio_threshold", 3.8)),
        semantic_peak_min_max = float(get("semantic_peak_min_max", 0.015)),
        semantic_drift_p90_exempt = float(get("semantic_drift_p90_exempt", 0.005)),
        semantic_drift_sparse_mid_count = int(get("semantic_drift_sparse_mid_count", 2)),

        semantic_patch_enabled = bool(get("semantic_patch_enabled", True)),
        semantic_patch_grid = int(get("semantic_patch_grid", 8)),
        semantic_patch_hot_threshold = float(get("semantic_patch_hot_threshold", 0.015)),
        semantic_patch_min_hot_patches = int(get("semantic_patch_min_hot_patches", 12)),
        semantic_patch_center_margin = float(get("semantic_patch_center_margin", 0.4)),
        semantic_patch_edge_margin = float(get("semantic_patch_edge_margin", 0.6)),
        semantic_patch_corner_merge_frac = float(get("semantic_patch_corner_merge_frac", 0.55)),
        semantic_patch_center_split_frac = float(get("semantic_patch_center_split_frac", 0.45)),
        semantic_patch_persistent_frame_frac = float(get("semantic_patch_persistent_frame_frac", 0.5)),
        semantic_patch_persistent_min = int(get("semantic_patch_persistent_min", 2)),
        semantic_patch_persistent_max = int(get("semantic_patch_persistent_max", 8)),
        semantic_patch_persistent_corner_min = float(get("semantic_patch_persistent_corner_min", 0.8)),
        semantic_patch_weak_center_max = float(get("semantic_patch_weak_center_max", 0.12)),
        semantic_patch_weak_hot_ratio_max = float(get("semantic_patch_weak_hot_ratio_max", 0.10)),
        semantic_patch_heavy_persistent_min = int(get("semantic_patch_heavy_persistent_min", 10)),
        semantic_patch_heavy_hot_ratio_min = float(get("semantic_patch_heavy_hot_ratio_min", 0.20)),
        semantic_patch_heavy_pers_corner_max = float(get("semantic_patch_heavy_pers_corner_max", 0.85)),
        semantic_patch_heavy_min_ratio = float(get("semantic_patch_heavy_min_ratio", 2.5)),
        semantic_patch_center_persistent_corner_max = float(get("semantic_patch_center_persistent_corner_max", 0.20)),
        semantic_patch_center_persistent_total_corner_max = float(get("semantic_patch_center_persistent_total_corner_max", 0.25)),
        semantic_patch_drift_exempt_center_persistent_blocks = bool(get("semantic_patch_drift_exempt_center_persistent_blocks", True)),
        semantic_patch_anim_hot_ratio_min = float(get("semantic_patch_anim_hot_ratio_min", 0.15)),
        semantic_patch_anim_corner_min = float(get("semantic_patch_anim_corner_min", 0.65)),
        semantic_patch_anim_center_max = float(get("semantic_patch_anim_center_max", 0.10)),
        semantic_patch_anim_max_override_factor = float(get("semantic_patch_anim_max_override_factor", 2.5)),
        semantic_patch_ratio_rescue_corner_min = float(get("semantic_patch_ratio_rescue_corner_min", 0.60)),
        semantic_patch_ratio_rescue_center_max = float(get("semantic_patch_ratio_rescue_center_max", 0.15)),
        semantic_patch_ratio_rescue_mean_max = float(get("semantic_patch_ratio_rescue_mean_max", 0.010)),
        semantic_patch_max_rescue_corner_min = float(get("semantic_patch_max_rescue_corner_min", 0.75)),
        semantic_patch_max_rescue_center_max = float(get("semantic_patch_max_rescue_center_max", 0.10)),
        semantic_patch_max_rescue_mean_max = float(get("semantic_patch_max_rescue_mean_max", 0.010)),
        semantic_patch_center_mask_enabled = bool(get("semantic_patch_center_mask_enabled", True)),
        semantic_patch_center_mask_inner = int(get("semantic_patch_center_mask_inner", 4)),
        semantic_patch_center_mask_mean_max = float(get("semantic_patch_center_mask_mean_max", 0.006)),
        semantic_patch_center_mask_max_max = float(get("semantic_patch_center_mask_max_max", 0.025)),
        semantic_patch_center_mask_hot_ratio_max = float(get("semantic_patch_center_mask_hot_ratio_max", 0.06)),
        semantic_patch_center_mask_relaxed_enabled = bool(get("semantic_patch_center_mask_relaxed_enabled", True)),
        semantic_patch_center_mask_relaxed_hot_ratio_max = float(get("semantic_patch_center_mask_relaxed_hot_ratio_max", 0.14)),
        semantic_patch_center_mask_relaxed_dom_q_max = float(get("semantic_patch_center_mask_relaxed_dom_q_max", 0.30)),
        semantic_patch_center_mask_relaxed_corner_min = float(get("semantic_patch_center_mask_relaxed_corner_min", 0.85)),
        semantic_patch_center_mask_relaxed_center_max = float(get("semantic_patch_center_mask_relaxed_center_max", 0.05)),
        semantic_patch_center_mask_relaxed_mean_max = float(get("semantic_patch_center_mask_relaxed_mean_max", 0.0065)),
        semantic_patch_center_mask_relaxed_p90_max = float(get("semantic_patch_center_mask_relaxed_p90_max", 0.010)),
        semantic_patch_center_mask_relaxed_max_max = float(get("semantic_patch_center_mask_relaxed_max_max", 0.060)),

        llm_enabled = bool(get("llm_enabled", False)),
        llm_model_name = str(get("llm_model_name", "Qwen/Qwen3-8B")),
        llm_device = str(get("llm_device", "auto")),
        llm_quantization = str(get("llm_quantization", "4bit")),
        llm_max_new_tokens = int(get("llm_max_new_tokens", 32)),
        llm_max_input_tokens = int(get("llm_max_input_tokens", 1024)),
        llm_enable_thinking = bool(get("llm_enable_thinking", False)),
        llm_gray_color_max_excess = float(get("llm_gray_color_max_excess", 0.06)),
        llm_gray_semantic_mean_max_excess = float(get("llm_gray_semantic_mean_max_excess", 0.008)),
        llm_gray_semantic_max_max_excess = float(get("llm_gray_semantic_max_max_excess", 0.02)),
        llm_gray_semantic_peak_max_excess = float(get("llm_gray_semantic_peak_max_excess", 1.5)),
        llm_intercept_on_patch_uncertain = bool(get("llm_intercept_on_patch_uncertain", True)),
        llm_intercept_pixel_clean_mean = float(get("llm_intercept_pixel_clean_mean", 0.010)),
        llm_intercept_pixel_clean_max = float(get("llm_intercept_pixel_clean_max", 0.030)),
        llm_rescue_enabled = bool(get("llm_rescue_enabled", False)),
        llm_skip_if_no_meta = bool(get("llm_skip_if_no_meta", True)),

        force_merge_pairs = list(get("force_merge_pairs", []) or []),
        force_split_pairs = list(get("force_split_pairs", []) or []),

        max_workers_stage1 = int(get("max_workers_stage1", 8)),
        max_workers_stage2 = int(get("max_workers_stage2", 6)),
        ffprobe_timeout = int(get("ffprobe_timeout", 60)),
        ffmpeg_timeout  = int(get("ffmpeg_timeout", 60)),
        fpcalc_timeout  = int(get("fpcalc_timeout", 60)),
        log_file        = get("log_file", None),
        progress        = bool(get("progress", True)),
    )

def main():
    ap = argparse.ArgumentParser(
        description="Wallpaper Engine 重复视频检测（管线并行 + pHash 模糊匹配）"
    )
    ap.add_argument("-c", "--config", type=Path, default=None, help="config.toml")
    ap.add_argument("--verbose", action="store_true", help="DEBUG 级别日志")
    ap.add_argument("--trace", action="store_true", help="打印外部命令与耗时/首行错误")
    ap.add_argument("--no-progress", action="store_true", help="关闭进度条显示")
    args = ap.parse_args()

    cfg = load_config(args.config)
    cfg.verbose = bool(args.verbose)
    cfg.trace   = bool(args.trace)
    if args.no_progress:
        cfg.progress = False

    # 空字符串视为不写日志文件
    if cfg.log_file is not None and str(cfg.log_file).strip() == "":
        cfg.log_file = None

    # 输出目录先创建，用来放 state/cache
    root = Path(cfg.workshop_root).resolve()
    out_dir = Path(cfg.output_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    myprojects_root: Optional[Path] = None
    if cfg.include_myprojects:
        we_dir = str(cfg.we_install_dir or "").strip()
        if we_dir:
            mp = Path(we_dir).resolve() / "projects" / "myprojects"
            if mp.is_dir():
                myprojects_root = mp
                LOGGER.info("[INFO] 同时扫描 myprojects: %s", mp)
            else:
                LOGGER.info("[INFO] myprojects 目录不存在，跳过：%s", mp)
        else:
            LOGGER.info("[INFO] 已开启 include_myprojects 但未配置 we_install_dir，跳过 myprojects")

    # 运行状态：决定 log 文件是重建还是追加
    state_path = out_dir / "we_dedup_state.json"
    prev_status = None
    if state_path.exists():
        try:
            prev = json.loads(state_path.read_text(encoding="utf-8"))
            prev_status = prev.get("status")
        except Exception:
            prev_status = None
    file_mode = "w" if prev_status == "completed" else "a"

    setup_logging(logging.DEBUG if cfg.verbose else logging.INFO, cfg.log_file, file_mode=file_mode)

    # 初始化缓存：记住每个文件的时长/视觉/音频特征，再次跑只计算新文件，用缓存参与比对。
    # 注意：缓存采取"永久保留"策略——取消订阅删除文件后不清理缓存条目，
    #       便于后续重新订阅/下载时直接命中（按 path+size 复用所有特征）。
    global _CACHE
    try:
        _CACHE = SigCache(out_dir / "we_dedup_cache.sqlite3")
        LOGGER.info("[cache] 特征值缓存已加载（永久保留策略：已删文件的特征保留在库中，重新下载时直接命中）。")
    except Exception as e:
        _CACHE = None
        LOGGER.warning("[cache] 初始化失败（将不启用断点续跑）：%s", e)

    # 写入 running 状态（如果异常退出，下次会自动当作未完成并续跑）
    try:
        state_path.write_text(json.dumps({
            "status": "running",
            "start_ts": int(time.time()),
            "pid": os.getpid(),
            "log_file": cfg.log_file,
            "log_mode": file_mode,
        }, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass

    LOGGER.info(
        "[CFG] phash_distance_threshold=%.2f（数值越小越严格），tm_cap=%.2f，trim_ratio=%.2f，bimodal_gap_cap=%.2f",
        cfg.phash_distance_threshold, cfg.phash_trimmed_mean_cap,
        cfg.phash_trim_ratio, cfg.phash_bimodal_gap_cap,
    )

    exit_code = 0
    try:
        LOGGER.info("[INFO] Scanning workshop: %s", root)
        items_map = find_items(root, myprojects_root)
        LOGGER.info("[INFO] Found %d items with candidate video files", len(items_map))

        # 新主流程：阶段1 + 阶段2 管线并行（带缓存自动跳过）
        filesigs = stage1_and_stage2_pipelined(items_map, cfg)

        # 过滤可参与最终分组的文件
        eligible: List[FileSig] = []
        for fs in filesigs:
            if cfg.require_both_signatures:
                if fs.duration_bucket and fs.phash_parts and fs.audio_fp_digest:
                    eligible.append(fs)
            else:
                if fs.duration_bucket and fs.phash_parts:
                    eligible.append(fs)
        LOGGER.info("[INFO] 可参与最终分组的文件：%d / %d（候选）", len(eligible), len(filesigs))

        if not eligible:
            LOGGER.info("[INFO] 无可用签名文件，结束。")
            return

        # 粗分桶：按“时长分桶 +（可选）音频指纹”
        # 若 cfg.duration_rounding 变化，根据当前配置重算一次 bucket（不受旧缓存值约束）。
        def bucket_key(fs: FileSig):
            bk = nearest_bucket(fs.duration, cfg.duration_rounding) or fs.duration_bucket
            if cfg.require_both_signatures:
                return (bk, fs.audio_fp_digest)
            else:
                return (bk,)

        bucket_groups: Dict[Tuple, List[FileSig]] = defaultdict(list)
        all_file_sigs_by_item: Dict[str, List[FileSig]] = defaultdict(list)
        for fs in eligible:
            bucket_groups[bucket_key(fs)].append(fs)
            all_file_sigs_by_item[fs.item_id].append(fs)

        LOGGER.info(
            "[INFO] 粗分桶数量：%d（按时长%s，跨桶容差=%.2fs）",
            len(bucket_groups),
            "+音频" if cfg.require_both_signatures else "",
            float(cfg.duration_cross_bucket_tolerance),
        )

        def _run_cluster(flist: List[FileSig]) -> List[List[FileSig]]:
            return cluster_bucket_by_phash(
                flist,
                cfg.phash_distance_threshold,
                tm_cap=cfg.phash_trimmed_mean_cap,
                trim_ratio=cfg.phash_trim_ratio,
                bimodal_gap_cap=cfg.phash_bimodal_gap_cap,
                color_distance_threshold=cfg.color_distance_threshold,
                semantic_distance_threshold=cfg.semantic_distance_threshold,
                semantic_max_threshold=cfg.semantic_max_threshold,
                semantic_peak_ratio_threshold=cfg.semantic_peak_ratio_threshold,
                semantic_peak_min_max=cfg.semantic_peak_min_max,
                semantic_drift_p90_exempt=cfg.semantic_drift_p90_exempt,
                semantic_drift_sparse_mid_count=cfg.semantic_drift_sparse_mid_count,
                semantic_patch_enabled=cfg.semantic_patch_enabled,
                semantic_patch_grid=cfg.semantic_patch_grid,
                semantic_patch_hot_threshold=cfg.semantic_patch_hot_threshold,
                semantic_patch_min_hot_patches=cfg.semantic_patch_min_hot_patches,
                semantic_patch_center_margin=cfg.semantic_patch_center_margin,
                semantic_patch_edge_margin=cfg.semantic_patch_edge_margin,
                semantic_patch_corner_merge_frac=cfg.semantic_patch_corner_merge_frac,
                semantic_patch_center_split_frac=cfg.semantic_patch_center_split_frac,
                semantic_patch_persistent_frame_frac=cfg.semantic_patch_persistent_frame_frac,
                semantic_patch_persistent_min=cfg.semantic_patch_persistent_min,
                semantic_patch_persistent_max=cfg.semantic_patch_persistent_max,
                semantic_patch_persistent_corner_min=cfg.semantic_patch_persistent_corner_min,
                semantic_patch_weak_center_max=cfg.semantic_patch_weak_center_max,
                semantic_patch_weak_hot_ratio_max=cfg.semantic_patch_weak_hot_ratio_max,
                semantic_patch_heavy_persistent_min=cfg.semantic_patch_heavy_persistent_min,
                semantic_patch_heavy_hot_ratio_min=cfg.semantic_patch_heavy_hot_ratio_min,
                semantic_patch_heavy_pers_corner_max=cfg.semantic_patch_heavy_pers_corner_max,
                semantic_patch_heavy_min_ratio=cfg.semantic_patch_heavy_min_ratio,
                semantic_patch_center_persistent_corner_max=cfg.semantic_patch_center_persistent_corner_max,
                semantic_patch_center_persistent_total_corner_max=cfg.semantic_patch_center_persistent_total_corner_max,
                semantic_patch_drift_exempt_center_persistent_blocks=cfg.semantic_patch_drift_exempt_center_persistent_blocks,
                semantic_patch_anim_hot_ratio_min=cfg.semantic_patch_anim_hot_ratio_min,
                semantic_patch_anim_corner_min=cfg.semantic_patch_anim_corner_min,
                semantic_patch_anim_center_max=cfg.semantic_patch_anim_center_max,
                semantic_patch_anim_max_override_factor=cfg.semantic_patch_anim_max_override_factor,
                semantic_patch_ratio_rescue_corner_min=cfg.semantic_patch_ratio_rescue_corner_min,
                semantic_patch_ratio_rescue_center_max=cfg.semantic_patch_ratio_rescue_center_max,
                semantic_patch_ratio_rescue_mean_max=cfg.semantic_patch_ratio_rescue_mean_max,
                semantic_patch_max_rescue_corner_min=cfg.semantic_patch_max_rescue_corner_min,
                semantic_patch_max_rescue_center_max=cfg.semantic_patch_max_rescue_center_max,
                semantic_patch_max_rescue_mean_max=cfg.semantic_patch_max_rescue_mean_max,
                semantic_patch_center_mask_enabled=cfg.semantic_patch_center_mask_enabled,
                semantic_patch_center_mask_inner=cfg.semantic_patch_center_mask_inner,
                semantic_patch_center_mask_mean_max=cfg.semantic_patch_center_mask_mean_max,
                semantic_patch_center_mask_max_max=cfg.semantic_patch_center_mask_max_max,
                semantic_patch_center_mask_hot_ratio_max=cfg.semantic_patch_center_mask_hot_ratio_max,
                semantic_patch_center_mask_relaxed_enabled=cfg.semantic_patch_center_mask_relaxed_enabled,
                semantic_patch_center_mask_relaxed_hot_ratio_max=cfg.semantic_patch_center_mask_relaxed_hot_ratio_max,
                semantic_patch_center_mask_relaxed_dom_q_max=cfg.semantic_patch_center_mask_relaxed_dom_q_max,
                semantic_patch_center_mask_relaxed_corner_min=cfg.semantic_patch_center_mask_relaxed_corner_min,
                semantic_patch_center_mask_relaxed_center_max=cfg.semantic_patch_center_mask_relaxed_center_max,
                semantic_patch_center_mask_relaxed_mean_max=cfg.semantic_patch_center_mask_relaxed_mean_max,
                semantic_patch_center_mask_relaxed_p90_max=cfg.semantic_patch_center_mask_relaxed_p90_max,
                semantic_patch_center_mask_relaxed_max_max=cfg.semantic_patch_center_mask_relaxed_max_max,
                audio_merge_override_color=cfg.audio_merge_override_color,
                audio_merge_threshold=cfg.audio_merge_threshold,
                llm_enabled=cfg.llm_enabled,
                llm_cache=_CACHE,
                llm_model_name=cfg.llm_model_name,
                llm_model_cache_dir=cfg.model_cache_dir,
                llm_device=cfg.llm_device,
                llm_quantization=cfg.llm_quantization,
                llm_max_new_tokens=cfg.llm_max_new_tokens,
                llm_max_input_tokens=cfg.llm_max_input_tokens,
                llm_enable_thinking=cfg.llm_enable_thinking,
                llm_gray_color_max_excess=cfg.llm_gray_color_max_excess,
                llm_gray_semantic_mean_max_excess=cfg.llm_gray_semantic_mean_max_excess,
                llm_gray_semantic_max_max_excess=cfg.llm_gray_semantic_max_max_excess,
                llm_gray_semantic_peak_max_excess=cfg.llm_gray_semantic_peak_max_excess,
                llm_intercept_on_patch_uncertain=cfg.llm_intercept_on_patch_uncertain,
                llm_intercept_pixel_clean_mean=cfg.llm_intercept_pixel_clean_mean,
                llm_intercept_pixel_clean_max=cfg.llm_intercept_pixel_clean_max,
                llm_rescue_enabled=cfg.llm_rescue_enabled,
                llm_skip_if_no_meta=cfg.llm_skip_if_no_meta,
            )

        # 在每个粗桶内用 phash_parts 做模糊聚类
        all_duplicate_groups: List[List[FileSig]] = []
        for bkey, flist in bucket_groups.items():
            if len(flist) < 2:
                continue
            clusters = _run_cluster(flist)
            if not clusters:
                continue
            all_duplicate_groups.extend(clusters)
            LOGGER.info("[dup-bucket] 粗桶 %s 内 fuzzy 子组数=%d", bkey, len(clusters))

        # —— 跨相邻桶比较：救回"同源视频时长跨桶边界"的对 ——
        # 例如 29.6s / 30.2s 在整秒分桶下会分到 30 和 30（都四舍五入到 30），
        # 但 30.4s / 31.1s 分到 30 和 31——之前永远不比较。此步合并相邻桶临时 cluster。
        # 产生的 cluster 会跟原单桶 cluster 部分重叠（跨桶对的双端都会在单桶里也有伙伴），
        # 最后用 `_merge_overlapping_clusters` 通过 item_id 并查集去重合并。
        if cfg.duration_cross_bucket_tolerance > 0.0 and len(bucket_groups) >= 2:
            def _bk_to_float(bk_str) -> float:
                try:
                    return float(bk_str)
                except (ValueError, TypeError):
                    return 0.0

            sorted_bkeys = sorted(bucket_groups.keys(), key=lambda k: _bk_to_float(k[0]))
            cross_pair_count = 0
            cross_cluster_count = 0
            for idx in range(len(sorted_bkeys) - 1):
                k1 = sorted_bkeys[idx]
                k2 = sorted_bkeys[idx + 1]
                d1 = _bk_to_float(k1[0])
                d2 = _bk_to_float(k2[0])
                if d2 - d1 > float(cfg.duration_cross_bucket_tolerance):
                    continue
                # 要求 audio_fp 签名时，相邻桶的音频指纹必须匹配才交叉
                if cfg.require_both_signatures:
                    if len(k1) < 2 or len(k2) < 2 or k1[1] != k2[1]:
                        continue
                merged_flist = bucket_groups[k1] + bucket_groups[k2]
                if len(merged_flist) < 2:
                    continue
                cross_clusters = _run_cluster(merged_flist)
                if not cross_clusters:
                    continue
                all_duplicate_groups.extend(cross_clusters)
                cross_pair_count += 1
                cross_cluster_count += len(cross_clusters)
            if cross_pair_count:
                LOGGER.info(
                    "[dup-cross] 相邻桶交叉比较：处理桶对=%d，产生临时子组=%d（稍后合并重叠）",
                    cross_pair_count, cross_cluster_count,
                )

            # 合并"同一 item 同时出现在多个 cluster"的重叠子组（跨桶交叉必然导致重叠）
            all_duplicate_groups = _merge_overlapping_clusters(all_duplicate_groups)
            LOGGER.info(
                "[dup-cross] 合并重叠后最终子组数=%d",
                len(all_duplicate_groups),
            )

        # —— 人工覆写：force_split / force_merge ——
        # 处理算法仍无法确信的极少数边缘对。按 wid 对应 item_id 匹配。
        if cfg.force_split_pairs or cfg.force_merge_pairs:
            all_duplicate_groups = _apply_force_overrides(
                all_duplicate_groups,
                cfg.force_merge_pairs,
                cfg.force_split_pairs,
                all_file_sigs_by_item=all_file_sigs_by_item,
            )

        # 组织导出：对每个 fuzzy 组，合并到“不同 item”，并按该 item 命中的最大文件大小降序。
        # 同一 item 可能在工坊/myprojects 下有多份；按"文件大小最大的那份"选代表，url 与 folder 同步对齐。
        duplicate_groups_urls: List[List[str]] = []
        duplicate_groups_folders: List[List[str]] = []
        kept_groups = 0
        for group in all_duplicate_groups:
            item_to_bestsize: Dict[str, int] = {}
            item_to_url: Dict[str, str] = {}
            item_to_folder: Dict[str, str] = {}
            for fs in group:
                if fs.item_id not in item_to_bestsize or fs.size > item_to_bestsize[fs.item_id]:
                    item_to_bestsize[fs.item_id] = fs.size
                    item_to_url[fs.item_id] = fs.url
                    # 视频文件所在的文件夹路径；workshop 是 `…\431960\<id>\`，myprojects 是 `…\myprojects\<子文件夹>\`
                    try:
                        item_to_folder[fs.item_id] = str(fs.path.parent.resolve())
                    except Exception:
                        item_to_folder[fs.item_id] = str(fs.path.parent)
            if len(item_to_bestsize) <= 1:
                continue
            # 首列 = 取消订阅流程中保留的那份。按文件大小降序；size 相等时
            # myprojects（item_id 以 "mp:" 开头）优先排前，这样同分辨率同码率的
            # workshop+myprojects 会把 myprojects 作为保留项，被退订的永远是
            # 还没归档的 workshop 那份。
            def _dup_sort_key(kv: Tuple[str, int]) -> Tuple[int, int]:
                iid, sz = kv
                is_mp = 1 if iid.startswith("mp:") else 0
                return (sz, is_mp)
            ordered = sorted(item_to_bestsize.items(), key=_dup_sort_key, reverse=True)
            urls = [item_to_url[iid] for iid, _ in ordered]
            folders = [item_to_folder[iid] for iid, _ in ordered]
            duplicate_groups_urls.append(urls)
            duplicate_groups_folders.append(folders)
            kept_groups += 1
            LOGGER.info(
                "[dup] fuzzy 组 #%d → items=%s",
                kept_groups,
                [u.rsplit('=', 1)[-1] for u in urls]
            )

        # 导出：两份 xlsx
        #   duplicates_{ts}.xlsx         —— 创意工坊链接版（bulk_unsub_controller 取消订阅用，保持文件名 glob 不变）
        #   duplicate_paths_{ts}.xlsx    —— 文件夹路径版（方便在资源管理器手工处理；有意与 duplicates_* glob 区分开，避免被误选进取消订阅）
        ts = time.strftime("%Y%m%d_%H%M%S")
        out_xlsx_urls = out_dir / f"duplicates_{ts}.xlsx"
        out_xlsx_paths = out_dir / f"duplicate_paths_{ts}.xlsx"
        if duplicate_groups_urls:
            export_xlsx(duplicate_groups_urls, out_xlsx_urls,
                        sheet_title="duplicates", col_prefix="链接", hyperlink=True)
            export_xlsx(duplicate_groups_folders, out_xlsx_paths,
                        sheet_title="folders", col_prefix="路径", hyperlink=True)
            LOGGER.info("[OUT] XLSX  (URL)   : %s", out_xlsx_urls)
            LOGGER.info("[OUT] XLSX  (PATH)  : %s", out_xlsx_paths)
        else:
            LOGGER.info("[INFO] 未发现重复组（在当前 fuzzy 条件下）")

        LOGGER.info("[DONE] 任务完成，重复组数：%d", kept_groups)
    except KeyboardInterrupt:
        exit_code = 130
        LOGGER.warning("[EXIT] 用户中断（下次将从未完成部分继续）")
    except Exception as e:
        exit_code = 1
        LOGGER.exception("[EXIT] 异常退出（下次将从未完成部分继续）：%s", e)
    finally:
        # 写入最终状态
        try:
            status = "completed" if exit_code == 0 else "incomplete"
            state_path.write_text(json.dumps({
                "status": status,
                "end_ts": int(time.time()),
                "pid": os.getpid(),
                "exit_code": exit_code,
                "log_file": cfg.log_file,
            }, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass
        if _CACHE:
            try:
                _CACHE.close()
            except Exception:
                pass
            _CACHE = None
        # 关闭日志 handler，释放文件占用
        for h in list(LOGGER.handlers):
            try:
                h.close()
            except Exception:
                pass

    if exit_code != 0:
        raise SystemExit(exit_code)

if __name__ == "__main__":
    main()