# bulk_unsub_pool_controller.py
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse, os, shutil, time, subprocess, webbrowser, json, threading
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path
from typing import List, Deque, Optional, Set, Tuple
from urllib.parse import urlparse, unquote, urlunparse, parse_qs, urlencode
from urllib.request import url2pathname
from collections import deque, Counter
import re, requests
from openpyxl import load_workbook

REQ_HEADERS = {"User-Agent": "Mozilla/5.0 Chrome/124 Safari/537.36"}
REQ_TIMEOUT = 15
APP_RE = re.compile(r'/app/(\d+)\b', re.IGNORECASE)

# ========== 简单队列 ==========
class UrlQueue:
    def __init__(self, urls: List[str], single_page_mode: bool = False):
        self.lock = threading.Lock()
        self.urls = deque(urls)
        self.assigned = 0
        self.done = 0
        self.ok = 0
        self.fail = 0
        self.single_page_mode = single_page_mode
    def pop(self) -> str:
        with self.lock:
            if not self.urls: return ""
            self.assigned += 1
            return self.urls.popleft()
    def stats(self):
        with self.lock:
            return dict(left=len(self.urls), assigned=self.assigned, done=self.done, ok=self.ok, fail=self.fail)
    def mark_done(self, ok: bool|None = None):
        with self.lock:
            self.done += 1
            if ok is True:
                self.ok += 1
            elif ok is False:
                self.fail += 1

QUEUE: UrlQueue|None = None

# ========== 回调服务（CORS） ==========
def _cors(h):
    h.send_header("Access-Control-Allow-Origin", "*")
    h.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
    h.send_header("Access-Control-Allow-Headers", "content-type")
    h.send_header("Access-Control-Max-Age", "86400")

class Handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        path = self.path.rstrip('/')
        if path not in ["/next", "/batch", "/stats", "/report"]:
            self.send_response(404); _cors(self); self.end_headers(); return
        self.send_response(204); _cors(self); self.end_headers()
    
    def do_GET(self):
        path = self.path.rstrip('/')
        # 获取统计信息
        if path == "/stats":
            stats = QUEUE.stats() if QUEUE else {"left": 0, "assigned": 0, "done": 0, "ok": 0, "fail": 0}
            self.send_response(200); _cors(self)
            self.send_header("Content-Type","application/json"); self.end_headers()
            self.wfile.write(json.dumps(stats).encode('utf-8'))
            return
        # 批量获取 ID（用于单页面模式）
        elif path.startswith("/batch"):
            # 解析批次大小
            query = parse_qs(urlparse(self.path).query)
            size = int(query.get('size', ['10'])[0])
            ids = []
            if QUEUE and QUEUE.single_page_mode:
                for _ in range(size):
                    url = QUEUE.pop()
                    if not url: break
                    wid = extract_workshop_id(url)
                    if wid: ids.append(wid)
            print(f"[BATCH] 返回 {len(ids)} 个 workshop ID")
            self.send_response(200); _cors(self)
            self.send_header("Content-Type","application/json"); self.end_headers()
            self.wfile.write(json.dumps({"ids": ids}).encode('utf-8'))
            return
        else:
            self.send_response(404); _cors(self); self.end_headers()
    
    def do_POST(self):
        path = self.path.rstrip('/')
        if path not in ["/next", "/report"]:
            self.send_response(404); _cors(self); self.end_headers(); return
        length = int(self.headers.get('Content-Length','0') or 0)
        raw = self.rfile.read(length) if length else b'{}'
        try:
            data = json.loads(raw.decode('utf-8','ignore'))
        except Exception:
            data = {}
        slot   = (data.get('slot') or '')
        wid    = (data.get('id')   or '')
        status = (data.get('status') or '')
        ok_flag = data.get('ok', None)

        # 推断成功/失败（允许前端直接传 ok=true/false；否则按 status 粗略判断）
        ok: bool|None = None
        if isinstance(ok_flag, bool):
            ok = ok_flag
        else:
            s = str(status or '').lower()
            if 'fail' in s or 'error' in s:
                ok = False
            elif s:
                ok = True

        if QUEUE:
            QUEUE.mark_done(ok)

        if path == "/report":
            # 单页面模式：只上报结果，不分配下一条，避免“吞队列”
            print(f"[REPORT] id={wid} status={status} ok={ok}")
            self.send_response(200); _cors(self); self.send_header("Content-Type","application/json"); self.end_headers()
            self.wfile.write(json.dumps({"ok": True}).encode('utf-8'))
            return

        # /next：多页面模式继续分配下一条
        nxt = QUEUE.pop() if QUEUE else ""
        if QUEUE and QUEUE.single_page_mode and nxt:
            nxt_id = extract_workshop_id(nxt)
            print(f"[NEXT] slot={slot} id={wid} status={status} -> {'ASSIGN ID ' + nxt_id if nxt_id else 'EMPTY'}")
            self.send_response(200); _cors(self); self.send_header("Content-Type","application/json"); self.end_headers()
            self.wfile.write(json.dumps({"id": nxt_id, "url": ""}).encode('utf-8'))
        else:
            print(f"[NEXT] slot={slot} id={wid} status={status} -> {'ASSIGN ' + nxt if nxt else 'EMPTY'}")
            self.send_response(200); _cors(self); self.send_header("Content-Type","application/json"); self.end_headers()
            self.wfile.write(json.dumps({"url": nxt}).encode('utf-8'))

def start_server(port: int):
    srv = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    t = threading.Thread(target=srv.serve_forever, daemon=True)
    t.start()
    return srv

# ========== 打开 URL ==========
def open_url(url: str):
    try:
        if os.name == "nt":
            try: os.startfile(url); return
            except Exception: pass
            try:
                subprocess.Popen(
                    ["powershell", "-NoProfile", "-WindowStyle", "Hidden",
                     "-Command", "Start-Process", url],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                    creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0)
                ); return
            except Exception: pass
            subprocess.Popen(f'start "" "{url}"', shell=True); return
        else:
            webbrowser.open_new_tab(url)
    except Exception:
        webbrowser.open_new_tab(url)

# ========== Excel & URL 工具 ==========

def _is_link_cell(s: str) -> bool:
    t = s.strip()
    return t.startswith("http") or t.lower().startswith("file:")


def _file_uri_to_path(uri: str) -> Optional[Path]:
    u = uri.strip()
    p = urlparse(u)
    if p.scheme != "file":
        return None
    try:
        raw = p.path or ""
        # Windows: file:///C:/x -> path /C:/x
        local = url2pathname(unquote(raw))
        return Path(local)
    except Exception:
        return None


def _parse_myprojects_video_path(cell: str) -> Optional[Path]:
    """将单元格解析为视频文件路径（file: 或含盘符的绝对路径）。"""
    s = (cell or "").strip()
    if not s:
        return None
    if s.lower().startswith("file:"):
        return _file_uri_to_path(s)
    # 裸 Windows/UNC 路径（筛重导出一般为 file:，兼容手动粘贴）
    if len(s) > 2 and s[1] == ":" and "myprojects" in s.replace("\\", "/").lower():
        try:
            return Path(s)
        except Exception:
            return None
    if s.startswith("\\\\") and "myprojects" in s.replace("\\", "/").lower():
        try:
            return Path(s)
        except Exception:
            return None
    return None


def _myprojects_item_root(video_path: Path) -> Optional[Path]:
    """.../projects/myprojects/<子文件夹>/xxx.mp4 -> .../myprojects/<子文件夹>"""
    try:
        p = video_path.resolve()
    except Exception:
        p = video_path
    parts = p.parts
    lower = [x.lower() for x in parts]
    for i, seg in enumerate(lower):
        if seg == "myprojects" and i + 1 < len(parts):
            return Path(*parts[: i + 2])
    return None


def _safe_rmtree_myprojects_item(root: Path, deleted: Set[str]) -> bool:
    """删除 myprojects 下的单个项目文件夹（整夹）。同一任务内去重。"""
    try:
        root = root.resolve()
    except Exception:
        pass
    key = str(root)
    if key in deleted:
        return True
    if not root.is_dir():
        print(f"[DELETE-SKIP] 不是目录或不存在: {root}")
        return False
    parts_lower = [p.lower() for p in root.parts]
    if "myprojects" not in parts_lower:
        print(f"[DELETE-SKIP] 路径不在 myprojects 下，拒绝删除: {root}")
        return False
    # 必须是 .../myprojects/<一层子目录>，防止误删整个 myprojects
    try:
        mp_idx = parts_lower.index("myprojects")
    except ValueError:
        return False
    if len(root.parts) != mp_idx + 2:
        print(f"[DELETE-SKIP] 只删除 myprojects 下直接子文件夹，跳过: {root}")
        return False
    try:
        shutil.rmtree(root)
        deleted.add(key)
        print(f"[DELETE] 已删除本地 myprojects 项目夹: {root}")
        return True
    except Exception as e:
        print(f"[DELETE-FAIL] {root}: {e}")
        return False


def read_urls_from_xlsx(xlsx_path: Path) -> Tuple[List[str], int]:
    """
    读取筛重 duplicates_*.xlsx：每行**第一个链接保留**（要留下的那份）；
    从第二列起的链接：myprojects 本地 file:/// 则**删除项目文件夹**，http Steam 链接进入取消订阅队列。
    返回 (steam_urls, deleted_folder_count)。
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    out: List[str] = []
    deleted_roots: Set[str] = set()
    delete_ops = 0
    it = ws.iter_rows(values_only=True)
    header = next(it, None)  # noqa: F841
    first = next(it, None)
    if not first:
        wb.close()
        return out, 0

    def first_link_col_idx(r):
        for i, cell in enumerate(r):
            if isinstance(cell, str) and _is_link_cell(cell):
                return i
        return None

    start_idx = first_link_col_idx(first)
    if start_idx is None:
        start_idx = 1  # 默认从 B 列（筛重 duplicates_*.xlsx）

    def collect_link_cells(row) -> List[str]:
        urls: List[str] = []
        for cell in row[start_idx:]:
            if isinstance(cell, str):
                u = cell.strip()
                if _is_link_cell(u):
                    urls.append(u)
        return urls

    def consume_row(row):
        nonlocal delete_ops
        links = collect_link_cells(row)
        if not links:
            return
        targets = links[1:]  # 首列始终保留（筛重已按最大文件排在第一列）
        for u in targets:
            vp = _parse_myprojects_video_path(u)
            if vp is not None and vp.exists():
                root = _myprojects_item_root(vp)
                if root and _safe_rmtree_myprojects_item(root, deleted_roots):
                    delete_ops += 1
                elif root is None:
                    print(f"[DELETE-SKIP] 无法定位 myprojects 项目目录: {u[:120]}")
            elif u.strip().lower().startswith("file:"):
                print(f"[DELETE-SKIP] file: 链接无法解析或文件已不存在: {u[:120]}")
            elif u.strip().startswith("http"):
                out.append(u.strip())

    consume_row(first)
    for row in it:
        if not row:
            continue
        consume_row(row)
    wb.close()
    return out, delete_ops

def add_or_update_query(url: str, key: str, value: str) -> str:
    p = urlparse(url); q = parse_qs(p.query); q[key] = [value]
    new_query = urlencode({k: v[-1] for k, v in q.items()})
    return urlunparse(p._replace(query=new_query))

def set_hash_flag(url: str, flag: str = "bulk_unsub=1") -> str:
    p = urlparse(url)
    return urlunparse(p._replace(fragment=flag))

def resolve_appid_once(workshop_id: str) -> str:
    url = f"https://steamcommunity.com/sharedfiles/filedetails/?id={workshop_id}"
    r = requests.get(url, headers=REQ_HEADERS, timeout=REQ_TIMEOUT)
    r.raise_for_status()
    html = r.text
    hrefs = re.findall(r'href="([^"]+)"', html, flags=re.IGNORECASE)
    hits = []
    for h in hrefs:
        m = APP_RE.search(h)
        if m: hits.append(m.group(1))
    if not hits:
        hits = APP_RE.findall(html)
    if not hits:
        return ""
    return Counter(hits).most_common(1)[0][0]

def resolve_appid(workshop_id: str, try_times: int = 2, cooldown_sec: float = 0.2) -> str:
    for _ in range(try_times):
        try:
            appid = resolve_appid_once(workshop_id)
            if appid: return appid
        except Exception:
            pass
        time.sleep(cooldown_sec)
    return ""

def extract_workshop_id(url: str) -> str:
    p = urlparse(url); q = parse_qs(p.query)
    return (q.get("id", [""])[0] or "").strip()

# ========== 主程序 ==========
def main():
    global QUEUE
    
    ap = argparse.ArgumentParser(description="Steam Workshop 批量取消订阅 控制器（固定池轮转，标签自拉取下一条）")
    ap.add_argument("--xlsx", type=Path, required=True)
    ap.add_argument("--batch-size", type=int, default=10, help="初始打开的标签数（固定池大小）")
    ap.add_argument("--add-appid", action="store_true", help="为每个链接补充 ?appid=XXXX")
    ap.add_argument("--notify-port", type=int, default=8787, help="本地回调端口（与脚本一致）")
    ap.add_argument("--single-page", action="store_true", help="单页面模式：在固定页面上批量取消订阅，避免触发速率限制")
    ap.add_argument("--single-page-url", type=str, default="", help="单页面模式使用的固定 URL（默认使用列表中的第一个）")
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"找不到文件：{args.xlsx}"); return
    raw, n_local_deleted = read_urls_from_xlsx(args.xlsx)
    if n_local_deleted:
        print(f"[LOCAL] 已直接删除 {n_local_deleted} 个 myprojects 项目文件夹（每行非首列）")
    if not raw:
        print("Excel 中无待取消订阅的 Steam 链接（本地 myprojects 已按上表处理）。")
        if n_local_deleted:
            return
        print("（也未执行任何本地删除。）")
        return

    # 去重
    seen=set(); urls=[]
    for u in raw:
        if u not in seen:
            seen.add(u); urls.append(u)

    # 预处理（可选 appid，统一 hash 触发）
    final = []
    for u in urls:
        if args.add_appid:
            wid = extract_workshop_id(u)
            if wid:
                appid = resolve_appid(wid)
                if appid:
                    u = add_or_update_query(u, "appid", appid)
        final.append(set_hash_flag(u, "bulk_unsub=1"))

    # 单页面模式
    if args.single_page:
        print("[MODE] 单页面模式：所有取消订阅操作将在固定页面上运行")
        QUEUE = UrlQueue(final, single_page_mode=True)
        
        # 确定要打开的固定页面
        # 默认改为“已订阅物品主页”，避免依赖某个具体创意工坊详情页
        # - /my/ 会自动指向当前登录账号
        # - appid=431960：Wallpaper Engine（如你需要全局订阅列表，可手动去掉该参数）
        default_subs_url = "https://steamcommunity.com/my/myworkshopfiles/?browsesort=mysubscriptions&browsefilter=mysubscriptions&appid=431960&p=1"
        if args.single_page_url:
            base_url = args.single_page_url
        else:
            base_url = default_subs_url
        
        # 确保 URL 带有 bulk_unsub=1 标记
        base_url = set_hash_flag(base_url, "bulk_unsub=1")
        
        # 启动回调服务
        srv = start_server(args.notify_port)
        print(f"[SERVER] http://127.0.0.1:{args.notify_port}")
        print(f"[SERVER] - /next   : 获取下一个 workshop ID")
        print(f"[SERVER] - /batch  : 批量获取 workshop ID")
        print(f"[SERVER] - /stats  : 查看队列统计")
        print(f"[QUEUE] total={len(final)} 个项目待取消订阅")
        
        # 打开固定数量的标签（避免打开太多）
        batch = min(args.batch_size, 5)  # 单页面模式最多打开 5 个标签
        print(f"[OPEN] 打开 {batch} 个固定页面标签...")
        for i in range(batch):
            open_url(base_url)
            print(f"[OPEN] {i+1}/{batch}: {base_url}")
            time.sleep(0.5)  # 稍微延迟，避免同时打开太多
        
        print("[RUNNING] 单页面模式运行中...（控制台会打印取消订阅进度）")
        
    # 多页面模式（原有逻辑）
    else:
        print("[MODE] 多页面模式：每个项目打开一个单独的页面")
        batch = max(1, args.batch_size)
        first = final[:batch]
        rest  = final[batch:]
        QUEUE = UrlQueue(rest, single_page_mode=False)

        # 启动回调服务
        srv = start_server(args.notify_port)
        print(f"[SERVER] http://127.0.0.1:{args.notify_port}/next  （标签完成后会来这里拉取下一条）")
        print(f"[QUEUE] total={len(final)}  pool={batch}  remaining_in_queue={len(rest)}")

        # 打开首批标签（只开 batch 个）
        opened = 0
        for u in first:
            open_url(u); opened += 1
            print(f"[OPEN] {opened}/{batch}: {u}")

        print("[RUNNING] 等待各标签完成后自行拉取下一条…（控制台会打印 NEXT 分配信息）")

    print("\n[提示] 打开浏览器控制台(F12)可以查看详细执行日志")
    print("[提示] 按 Ctrl+C 可以随时停止")

    try:
        last_done = 0
        check_interval = 0
        while True:
            time.sleep(1)
            check_interval += 1
            st = QUEUE.stats()
            
            # 每5秒打印一次进度
            if check_interval >= 5:
                check_interval = 0
                if st['done'] > last_done:
                    print(f"[进度] 已完成: {st['done']}（成功:{st.get('ok',0)} 失败:{st.get('fail',0)}） | 队列剩余: {st['left']} | 已分配: {st['assigned']}")
                    last_done = st['done']
            
            if st['left']==0 and st['assigned'] == st['done']:
                # 队列空了且所有任务都完成了
                print(f"\n[完成] 所有任务已处理完毕！")
                print(f"[统计] 总计处理: {st['done']} 个项目（成功:{st.get('ok',0)} 失败:{st.get('fail',0)}）")
                print(f"[提示] 请查看浏览器控制台确认实际成功/失败数量")
                break
    except KeyboardInterrupt:
        st = QUEUE.stats()
        print(f"\n[EXIT] 用户中断")
        print(f"[统计] 已完成: {st['done']}（成功:{st.get('ok',0)} 失败:{st.get('fail',0)}） | 剩余: {st['left']}")
    finally:
        srv.shutdown(); srv.server_close()

if __name__ == "__main__":
    main()